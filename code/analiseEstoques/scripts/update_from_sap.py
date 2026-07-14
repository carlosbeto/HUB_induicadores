#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
update_from_sap.py

Atualiza o projeto a partir do Excel do SAP (snapshot diário), 4MDG e Base Família Comercial.
- Lê o SAP (aba Exportação SAPUI5)
- Calcula totais (qtd_total, val_total)
- Grava no SQLite (fact_estoque_snapshot)
- Registra baseline mensal (primeiro dia útil) se ainda não existir
- Lê o 4MDG (abas Produto e Peça) e grava/atualiza dim_material_proc
- Lê Base família comercial e grava/atualiza dim_material_grupo:
    material + grupo + familia + subfamilia
- Lê base de custo unitário e grava/atualiza dim_material_custo
- (V3) Lê MB51 (entradas) e grava em fact_mb51_entradas
- Gera OUTPUT pronto para consumo (MAST + MASR + WEPV):
    - RESUMO
    - EVOLUCAO_MES
    - BASE_CONSUMO (completa, último snapshot)
    - TOP_MATERIAIS (TOP N configurável)
    - PRIORIDADE_OPERACIONAL_MAST (somente MAST)
    - ENTRADAS_MAST e ENTRADAS_MAST_DIA (somente MAST)

Uso:
  venv\\Scripts\\python update_from_sap.py --config "..\\config.json"
  venv\\Scripts\\python update_from_sap.py --config "..\\config.json" --snapshot-date 2025-12-01
"""

# -----------------------------------------
# IMPORTS E CONSTANTES
# -----------------------------------------

import argparse
import json
import re
import sqlite3
import pandas as pd
import numpy as np
import openpyxl
import shutil
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime, date, timedelta
from typing import Optional, List, Dict, Any

from analiseEstoques.etl.load_dim_material_db import carregar_dim_material

SAP_SHEET_DEFAULT = "Exportação SAPUI5"


# ============================================================
# Helpers (datas / textos / colunas / SQLite safety)
# ============================================================

def escrever_excel_com_fallback(
    paths: List[Path],
    resumo_df: pd.DataFrame,
    evo_df: pd.DataFrame,
    gerar_base: bool,
    base_full: pd.DataFrame,
    gerar_top: bool,
    top_df: pd.DataFrame,
) -> Path:
    """
    Tenta escrever o Excel em uma lista de caminhos (em ordem).
    Retorna o caminho que conseguiu salvar.
    Se todos falharem com PermissionError, relança a exceção.
    """
    last_err = None
    for p in paths:
        try:
            with pd.ExcelWriter(p, engine="openpyxl") as xw:
                resumo_df.to_excel(xw, index=False, sheet_name="RESUMO")
                evo_df.to_excel(xw, index=False, sheet_name="EVOLUCAO_MES")
                if gerar_base:
                    base_full.to_excel(xw, index=False, sheet_name="BASE_CONSUMO")
                if gerar_top:
                    top_df.to_excel(xw, index=False, sheet_name="TOP_MATERIAIS")
            return p
        except PermissionError as e:
            last_err = e
            continue

    # Se chegou aqui, nenhum caminho estava disponível
    raise last_err if last_err else PermissionError("Falha ao gravar Excel: caminhos indisponíveis.")


def ajustar_excel_visual(path: Path):
    """
    Ajusta visual do Excel:
    - Auto largura de colunas (com base no conteúdo)
    - Formata números (2 casas), quantidades como inteiro, percentuais como %
    """
    wb = openpyxl.load_workbook(path)

    def fmt_por_header(h: str) -> str:
        h = (h or "").lower()

        # Percentuais
        if "pct" in h or "percent" in h:
            return "0.00%"

        # Quantidades
        if "qtd" in h:
            return "#,##0"

        # Valores / custos
        if "val" in h or "valor" in h or "custo" in h:
            return "#,##0.00"

        # Scores/índices
        if "score" in h or "indice" in h:
            return "0.00"

        return "0.00"

    for ws in wb.worksheets:
        headers = {}
        for col in range(1, ws.max_column + 1):
            headers[col] = ws.cell(row=1, column=col).value

        for col in range(1, ws.max_column + 1):
            header = str(headers.get(col, "") or "")
            num_fmt = fmt_por_header(header)

            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                v = cell.value
                if isinstance(v, (int, float)) and v is not None:
                    cell.number_format = num_fmt

        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row in range(1, ws.max_row + 1):
                v = ws.cell(row=row, column=col).value
                if v is None:
                    continue
                s = str(v)
                if len(s) > max_len:
                    max_len = len(s)

            width = min(max(max_len + 2, 10), 45)
            ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(path)


def now_ts() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def to_date_str(d: date) -> str:
    return d.strftime("%Y-%m-%d")


def mes_ref(d: date) -> str:
    return d.strftime("%Y-%m")


def first_business_day(d: date) -> date:
    """Primeiro dia útil (seg-sex), sem considerar feriados."""
    first = d.replace(day=1)
    while first.weekday() >= 5:
        first += timedelta(days=1)
    return first


def pick_latest_xlsx(folder: Path) -> Path:
    """Pega o .xlsx mais recente (por data de modificação) dentro de uma pasta.
       Ignora arquivos temporários do Excel (~$...)."""
    if not folder.exists():
        raise FileNotFoundError(f"Pasta não encontrada: {folder}")

    files = [
        p for p in folder.glob("*.xlsx")
        if not p.name.startswith("~$")
    ]

    if not files:
        raise FileNotFoundError(f"Nenhum .xlsx válido encontrado em {folder} (ignorando ~$.xlsx).")

    return max(files, key=lambda p: p.stat().st_mtime)


def normalize_colname(c: str) -> str:
    """Normaliza cabeçalhos (remove múltiplos espaços)."""
    return re.sub(r"\s+", " ", str(c)).strip()


def find_col(cols: List[str], startswith: List[str]) -> Optional[str]:
    """Procura coluna por prefixo (case-insensitive)."""
    cols_norm = {normalize_colname(c): c for c in cols}
    for s in startswith:
        for cn, orig in cols_norm.items():
            if cn.lower().startswith(s.lower()):
                return orig
    return None


def as_text_series(s: pd.Series) -> pd.Series:
    """Força texto (evita notação científica, preserva zeros à esquerda)."""
    return s.astype("string").str.strip()


def coerce_num(s: pd.Series) -> pd.Series:
    """Converte para número, vira NaN se não der."""
    return pd.to_numeric(s, errors="coerce")


def to_sql(v):
    """
    Converte valores do pandas (pd.NA, NaN, NaT) para None (SQLite).
    SQLite não aceita NAType.
    """
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass

    if hasattr(v, "item"):
        try:
            return v.item()
        except Exception:
            pass
    return v


def sanitize_rows(rows: List[tuple]) -> List[tuple]:
    """Aplica to_sql em todas as células de todas as linhas."""
    return [tuple(to_sql(x) for x in r) for r in rows]


def norm_material_text(x) -> str:
    """
    Normaliza material SEMPRE como texto:
    - remove ".0" no final
    - strip
    - retorna string vazia se None/NA
    """
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    s = str(x).strip()
    s = re.sub(r"\.0$", "", s)
    return s


def print_step(msg: str):
    """Log simples no console (ajuda muito na depuração)."""
    print(f"[{now_ts()}] {msg}")


def atualizar_arquivo_atual(ret_path: Path, out_dir: Path, deposito: str):
    """
    Cria/sobrescreve o arquivo '...__ATUAL.xlsx' na mesma pasta do OUTPUT do depósito.
    """
    atual_path = out_dir / f"{deposito}_resumo__ATUAL.xlsx"

    try:
        shutil.copy2(ret_path, atual_path)
        print_step(f"ATUAL atualizado: {atual_path.name} (fonte: {Path(ret_path).name})")
    except PermissionError:
        aviso_path = out_dir / f"AVISO_ATUAL_BLOQUEADO_{deposito}.txt"
        aviso_msg = (
            "ATENÇÃO – ARQUIVO '__ATUAL.xlsx' BLOQUEADO\n\n"
            f"Data/Hora: {now_ts()}\n"
            f"Depósito: {deposito}\n\n"
            "O arquivo fixo '__ATUAL.xlsx' estava aberto no momento da atualização\n"
            "e não pôde ser sobrescrito.\n\n"
            "Arquivo fonte gerado (que está atualizado):\n"
            f"{ret_path}\n\n"
            "Ação:\n"
            "1) Feche o arquivo '__ATUAL.xlsx'\n"
            "2) Rode a atualização novamente (ou copie manualmente o arquivo fonte)\n"
        )
        aviso_path.write_text(aviso_msg, encoding="utf-8")
        print_step(f"ATENÇÃO: '__ATUAL.xlsx' bloqueado. Aviso criado: {aviso_path.name}")


def safe_ratio(num, den):
    if den in (None, 0) or pd.isna(den):
        return None
    if num is None or pd.isna(num):
        return None
    return num / den


def table_has_column(con: sqlite3.Connection, table: str, col: str) -> bool:
    cur = con.execute(f"PRAGMA table_info({table});")
    cols = [r[1] for r in cur.fetchall()]
    return col in cols


# ============================================================
# Loaders (SAP / MB51 / 4MDG / Família Comercial / Custo)
# ============================================================

def load_sap(path: Path, sheet: str = SAP_SHEET_DEFAULT) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet)
    df.columns = [normalize_colname(c) for c in df.columns]

    col_material = "Material"
    col_desc = "Descrição de material"
    col_deposito = "Depósito"
    col_q_livre = "Estoque de utilização livre"
    col_q_qc = "Estoque em controle de qualidade"
    col_q_bloq = "Estoque bloqueado"
    col_v_livre = "Valor do estoque de utilização livre"
    col_v_qc = "Valor do estoque no controle de qualidade"
    col_v_bloq = "Valor do estoque bloqueado"

    required = [
        col_material, col_desc, col_deposito,
        col_q_livre, col_q_qc, col_q_bloq,
        col_v_livre, col_v_qc, col_v_bloq
    ]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Layout SAP inesperado. Colunas ausentes: {missing}")

    out = pd.DataFrame({
        "material": as_text_series(df[col_material]).apply(norm_material_text),
        "descricao": df[col_desc].astype("string"),
        "deposito": df[col_deposito].astype("string").str.strip().str.upper(),
        "qtd_livre": coerce_num(df[col_q_livre]),
        "qtd_qc": coerce_num(df[col_q_qc]),
        "qtd_bloq": coerce_num(df[col_q_bloq]),
        "val_livre": coerce_num(df[col_v_livre]),
        "val_qc": coerce_num(df[col_v_qc]),
        "val_bloq": coerce_num(df[col_v_bloq]),
    })

    out["qtd_total"] = out[["qtd_livre", "qtd_qc", "qtd_bloq"]].sum(axis=1, skipna=True)
    out["val_total"] = out[["val_livre", "val_qc", "val_bloq"]].sum(axis=1, skipna=True)

    out = out.dropna(subset=["material", "deposito"])
    out = out[out["material"].astype(str).str.strip() != ""]
    out = out[out["deposito"].astype(str).str.strip() != ""]

    return out

def load_mb51(path: Path, sheet_name: str = "Data") -> pd.DataFrame:
    """
    Lê o relatório MB51 (aba 'Data') e devolve um DataFrame normalizado.

    Colunas usadas:
      - Material
      - Texto breve material -> descricao_mb51
      - Data de lançamento   -> data_lancamento
      - Depósito             -> deposito
      - Quantidade           -> quantidade
      - Ordem                -> ordem
      - Cód.débito/crédito   -> deb_cred (S/H)
      - Tipo de movimento    -> tipo_movimento
      - Documento do depósito -> doc_deposito
      - Doc.material         -> doc_material
      - Item doc.material    -> item_doc_material
    """
    df = pd.read_excel(path, sheet_name=sheet_name)
    df.columns = [normalize_colname(c) for c in df.columns]

    col_material = find_col(df.columns.tolist(), ["Material"])
    col_desc     = find_col(df.columns.tolist(), ["Texto breve material", "Texto breve"])
    col_data     = find_col(df.columns.tolist(), ["Data de lançamento", "Data lancamento", "Data de lancamento"])
    col_deposito = find_col(df.columns.tolist(), ["Depósito", "Deposito"])
    col_qtd      = find_col(df.columns.tolist(), ["Quantidade"])
    col_ordem    = find_col(df.columns.tolist(), ["Ordem"])
    col_debcred  = find_col(df.columns.tolist(), ["Cód.débito/crédito", "Cod.debito/credito", "Cod débito/crédito", "Debito/credito"])
    col_tipomov  = find_col(df.columns.tolist(), ["Tipo de movimento", "Tipo movimento"])
    col_docdep   = find_col(df.columns.tolist(), ["Documento do depósito", "Documento do deposito", "Doc do depósito", "Doc do deposito"])
    col_docmat   = find_col(df.columns.tolist(), ["Doc.material", "Doc material", "Documento material"])
    col_itemdoc  = find_col(df.columns.tolist(), ["Item doc.material", "Item doc material", "Item documento material"])

    required = [
        ("Material", col_material),
        ("Data de lançamento", col_data),
        ("Depósito", col_deposito),
        ("Quantidade", col_qtd),
        ("Ordem", col_ordem),
        ("Cód.débito/crédito", col_debcred),
    ]
    missing = [name for name, col in required if not col]
    if missing:
        raise ValueError(
            f"Layout MB51 inesperado. Colunas ausentes: {missing}. "
            f"Encontrado: {list(df.columns)}"
        )

    out = pd.DataFrame({
        "material": as_text_series(df[col_material]).apply(norm_material_text),
        "descricao_mb51": (df[col_desc].astype("string") if col_desc else pd.Series([pd.NA]*len(df))).astype("string"),
        "data_lancamento": pd.to_datetime(df[col_data], errors="coerce", dayfirst=True),
        "deposito": df[col_deposito].astype("string").str.strip().str.upper(),
        "quantidade": coerce_num(df[col_qtd]),
        "ordem": df[col_ordem].astype("string"),
        "deb_cred": df[col_debcred].astype("string").fillna("").str.strip().str.upper(),

        "tipo_movimento": df[col_tipomov].astype("string").fillna("").str.strip() if col_tipomov else "",
        "doc_deposito": df[col_docdep].astype("string").fillna("").str.strip() if col_docdep else "",
        "doc_material": df[col_docmat].astype("string").fillna("").str.strip() if col_docmat else "",
        "item_doc_material": df[col_itemdoc].astype("string").fillna("").str.strip() if col_itemdoc else "",
    })

    out["ordem"] = out["ordem"].fillna("").astype("string").str.strip()
    out["descricao_mb51"] = out["descricao_mb51"].fillna("").astype("string").str.strip()

    out = out.dropna(subset=["material", "deposito", "data_lancamento"])
    out = out[out["material"].astype(str).str.strip() != ""]
    out = out[out["deposito"].astype(str).str.strip() != ""]
    out = out[out["deb_cred"].isin(["S", "H"])]

    return out


def load_mdg(path: Path, sheets: List[str]) -> pd.DataFrame:
    all_rows = []
    for sh in sheets:
        try:
            df = pd.read_excel(path, sheet_name=sh)
        except Exception:
            continue

        df.columns = [normalize_colname(c) for c in df.columns]

        col_mat = find_col(df.columns.tolist(), ["Codigo_do_material", "Código_do_material", "Codigo do material", "Material"])
        col_desc = find_col(df.columns.tolist(), ["Descricao_Curta", "Descrição_Curta", "Descricao Curta"])
        col_linha = find_col(df.columns.tolist(), ["Linha_Unidade_de_Negocio", "Linha Unidade de Negocio"])
        col_garantia = find_col(df.columns.tolist(), ["Tempo_de_Garantia", "Tempo de Garantia"])
        col_proc = find_col(df.columns.tolist(), ["Processamento Material"])
        col_saida = find_col(df.columns.tolist(), ["Saída Manufatura Produto", "Saida Manufatura Produto"])
        col_cod_reman = find_col(df.columns.tolist(), ["Codigo_Remanufaturado", "Código_Remanufaturado"])
        col_obs = find_col(df.columns.tolist(), ["Observacao", "Observação"])

        if not col_mat:
            continue

        tmp = pd.DataFrame({
            "material": as_text_series(df[col_mat]).apply(norm_material_text),
            "descricao_curta": df[col_desc].astype("string") if col_desc else pd.NA,
            "linha_unid_negocio": df[col_linha].astype("string") if col_linha else pd.NA,
            "garantia_meses": coerce_num(df[col_garantia]) if col_garantia else pd.NA,
            "processamento": df[col_proc].astype("string") if col_proc else pd.NA,
            "saida_manufatura": df[col_saida].astype("string") if col_saida else pd.NA,
            "codigo_remanufaturado": as_text_series(df[col_cod_reman]).apply(norm_material_text) if col_cod_reman else pd.NA,
            "observacao": df[col_obs].astype("string") if col_obs else pd.NA,
            "source_sheet": sh
        })

        tmp = tmp.dropna(subset=["material"])
        tmp = tmp[tmp["material"].astype(str).str.strip() != ""]
        all_rows.append(tmp)

    if not all_rows:
        return pd.DataFrame(columns=["material", "processamento", "saida_manufatura", "source_sheet"])

    base = pd.concat(all_rows, ignore_index=True)

    for c in ["processamento", "saida_manufatura", "descricao_curta", "linha_unid_negocio", "observacao"]:
        if c in base.columns:
            base[c] = base[c].astype("string").str.strip()

    def score_row(r):
        s = 0
        proc = r.get("processamento")
        saida = r.get("saida_manufatura")
        if pd.notna(proc) and str(proc).strip() != "":
            s += 2
        if pd.notna(saida) and str(saida).strip() != "":
            s += 2
        if r.get("source_sheet") == "Produto":
            s += 0.5
        return s

    base["_score"] = base.apply(score_row, axis=1)
    base = base.sort_values(["material", "_score"], ascending=[True, False])
    base = base.drop_duplicates(subset=["material"], keep="first").drop(columns=["_score"])
    return base


def load_familia_base(
    path: Path,
    sheet: str,
    col_material: str,
    col_grupo: str,
    col_familia: str,
    col_subfamilia: str,
    col_area_negocio: str = "Área Negócio Centro Material"
) -> pd.DataFrame:
    """
    Lê a Base Família Comercial e normaliza as informações por material.

    Campos carregados:
    - material
    - grupo
    - familia
    - subfamilia
    - area_negocio  -> BU / unidade de negócio solicitada pela diretoria
    """

    df = pd.read_excel(path, sheet_name=sheet)
    df.columns = [normalize_colname(c) for c in df.columns]

    required = [
        col_material,
        col_grupo,
        col_familia,
        col_subfamilia,
        col_area_negocio
    ]

    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            "Base família: colunas não encontradas. "
            f"Faltando: {missing}. Encontrado: {list(df.columns)}"
        )

    out = pd.DataFrame({
        "material": as_text_series(df[col_material]).apply(norm_material_text),
        "grupo": df[col_grupo].astype("string").str.strip(),
        "familia": df[col_familia].astype("string").str.strip(),
        "subfamilia": df[col_subfamilia].astype("string").str.strip(),

        # Nova dimensão solicitada:
        # origem Excel: 'Área Negócio Centro Material'
        # nome no banco/app: area_negocio
        "area_negocio": df[col_area_negocio].astype("string").str.strip(),
    })

    out = out.dropna(subset=["material"])
    out = out[out["material"].astype(str).str.strip() != ""]

    out["grupo"] = out["grupo"].fillna("SEM_GRUPO")
    out["familia"] = out["familia"].fillna("SEM_FAMILIA")
    out["subfamilia"] = out["subfamilia"].fillna("SEM_SUBFAMILIA")
    out["area_negocio"] = out["area_negocio"].fillna("SEM_BU")

    out = out.drop_duplicates(subset=["material"], keep="last")
    return out


def load_custo_base(path: Path, sheet: str, col_material: str, col_custo: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet)
    df.columns = [normalize_colname(c) for c in df.columns]

    if col_material not in df.columns or col_custo not in df.columns:
        raise ValueError(
            "Base custo: colunas não encontradas. "
            f"Esperado: '{col_material}' e '{col_custo}'. "
            f"Encontrado: {list(df.columns)}"
        )

    out = pd.DataFrame({
        "material": as_text_series(df[col_material]).apply(norm_material_text),
        "custo_unit": coerce_num(df[col_custo])
    })

    out = out.dropna(subset=["material"])
    out = out[out["material"].astype(str).str.strip() != ""]
    out = out.drop_duplicates(subset=["material"], keep="last")
    return out


# ============================================================
# SQLite: criação (garantir) e UPSERTs
# ============================================================

def ensure_dim_material_custo(con: sqlite3.Connection):
    con.execute("""
        CREATE TABLE IF NOT EXISTS dim_material_custo (
            material TEXT PRIMARY KEY,
            custo_unit REAL,
            source_file TEXT,
            source_last_modified TEXT,
            load_ts TEXT
        );
    """)


def upsert_dim_material_custo(con: sqlite3.Connection, df: pd.DataFrame, source_file: str, source_last_modified: str):
    load_ts = now_ts()
    rows = []
    for _, r in df.iterrows():
        rows.append((
            norm_material_text(r.get("material")),
            float(r["custo_unit"]) if pd.notna(r.get("custo_unit")) else None,
            source_file,
            source_last_modified,
            load_ts
        ))
    rows = sanitize_rows(rows)

    con.executemany("""
        INSERT INTO dim_material_custo (material, custo_unit, source_file, source_last_modified, load_ts)
        VALUES (?,?,?,?,?)
        ON CONFLICT(material) DO UPDATE SET
            custo_unit=excluded.custo_unit,
            source_file=excluded.source_file,
            source_last_modified=excluded.source_last_modified,
            load_ts=excluded.load_ts;
    """, rows)


def ensure_dim_material_grupo(con: sqlite3.Connection):
    con.execute("""
        CREATE TABLE IF NOT EXISTS dim_material_grupo (
            material TEXT PRIMARY KEY,
            grupo TEXT,
            familia TEXT,
            subfamilia TEXT,
            area_negocio TEXT,
            source_file TEXT,
            source_last_modified TEXT,
            load_ts TEXT
        );
    """)

    cols = [r[1] for r in con.execute("PRAGMA table_info(dim_material_grupo);").fetchall()]

    if "familia" not in cols:
        con.execute("ALTER TABLE dim_material_grupo ADD COLUMN familia TEXT;")
    if "subfamilia" not in cols:
        con.execute("ALTER TABLE dim_material_grupo ADD COLUMN subfamilia TEXT;")
    if "area_negocio" not in cols:
        con.execute("ALTER TABLE dim_material_grupo ADD COLUMN area_negocio TEXT;")


def upsert_dim_material_grupo(con: sqlite3.Connection, df: pd.DataFrame, source_file: str, source_last_modified: str):
    ensure_dim_material_grupo(con)

    load_ts = now_ts()
    rows = []

    for _, r in df.iterrows():
        rows.append((
            norm_material_text(r.get("material")),
            to_sql(r.get("grupo")),
            to_sql(r.get("familia")),
            to_sql(r.get("subfamilia")),
            to_sql(r.get("area_negocio")),
            source_file,
            source_last_modified,
            load_ts
        ))

    rows = sanitize_rows(rows)

    con.executemany("""
        INSERT INTO dim_material_grupo (
            material, grupo, familia, subfamilia, area_negocio,
            source_file, source_last_modified, load_ts
        )
        VALUES (?,?,?,?,?,?,?,?)
        ON CONFLICT(material) DO UPDATE SET
            grupo=excluded.grupo,
            familia=excluded.familia,
            subfamilia=excluded.subfamilia,
            area_negocio=excluded.area_negocio,
            source_file=excluded.source_file,
            source_last_modified=excluded.source_last_modified,
            load_ts=excluded.load_ts;
    """, rows)


def ensure_dim_material_proc(con: sqlite3.Connection):
    con.execute("""
        CREATE TABLE IF NOT EXISTS dim_material_proc (
            material TEXT PRIMARY KEY,
            descricao_curta TEXT,
            linha_unid_negocio TEXT,
            garantia_meses REAL,
            processamento TEXT,
            saida_manufatura TEXT,
            codigo_remanufaturado TEXT,
            observacao TEXT,
            source_file TEXT,
            source_sheet TEXT,
            source_last_modified TEXT,
            load_ts TEXT
        );
    """)


def upsert_dim_proc(con: sqlite3.Connection, df: pd.DataFrame, source_file: str, source_sheet: str, source_last_modified: str):
    ensure_dim_material_proc(con)

    load_ts = now_ts()
    rows = []
    for _, r in df.iterrows():
        rows.append((
            norm_material_text(r.get("material")),
            r.get("descricao_curta"),
            r.get("linha_unid_negocio"),
            float(r["garantia_meses"]) if pd.notna(r.get("garantia_meses")) else None,
            r.get("processamento"),
            r.get("saida_manufatura"),
            r.get("codigo_remanufaturado"),
            r.get("observacao"),
            source_file,
            r.get("source_sheet") or source_sheet,
            source_last_modified,
            load_ts
        ))
    rows = sanitize_rows(rows)

    con.executemany("""
      INSERT INTO dim_material_proc (
        material, descricao_curta, linha_unid_negocio, garantia_meses,
        processamento, saida_manufatura, codigo_remanufaturado,
        observacao, source_file, source_sheet, source_last_modified, load_ts
      ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
      ON CONFLICT(material) DO UPDATE SET
        descricao_curta=excluded.descricao_curta,
        linha_unid_negocio=excluded.linha_unid_negocio,
        garantia_meses=excluded.garantia_meses,
        processamento=excluded.processamento,
        saida_manufatura=excluded.saida_manufatura,
        codigo_remanufaturado=excluded.codigo_remanufaturado,
        observacao=excluded.observacao,
        source_file=excluded.source_file,
        source_sheet=excluded.source_sheet,
        source_last_modified=excluded.source_last_modified,
        load_ts=excluded.load_ts;
    """, rows)


def ensure_fact_snapshot(con: sqlite3.Connection):
    """
    Garante a existência da tabela fato de snapshots de estoque.

    Tabela: fact_estoque_snapshot
    Granularidade (chave primária):
      - 1 linha por (snapshot_date, deposito, material)

    Por que essa tabela existe:
    - Representa o "estado do estoque" em um determinado dia (snapshot)
    - É a base para:
      * gráficos de evolução (SUM val_total / SUM qtd_total por dia)
      * análises por material e por depósito
      * comparação com baseline (KPIs mensais)

    Campos principais:
    - qtd_* e val_*: quantidades/valores em diferentes status (livre, qc, bloqueado, total)
    - source_file: qual arquivo SAP gerou aquele snapshot
    - load_ts: quando o ETL gravou (auditoria)
    """
    con.execute("""
        CREATE TABLE IF NOT EXISTS fact_estoque_snapshot (
            snapshot_date TEXT NOT NULL,     -- 'YYYY-MM-DD' (data do snapshot)
            deposito TEXT NOT NULL,          -- MAST/MASR/WEPV
            material TEXT NOT NULL,          -- código do material (normalizado como texto)
            descricao TEXT,                  -- texto do material (quando disponível)

            -- Quantidades por status
            qtd_livre REAL,
            qtd_qc REAL,
            qtd_bloq REAL,
            qtd_total REAL,

            -- Valores por status
            val_livre REAL,
            val_qc REAL,
            val_bloq REAL,
            val_total REAL,

            -- Auditoria de carga
            source_file TEXT,                -- nome do arquivo SAP usado
            load_ts TEXT,                    -- timestamp da carga no ETL

            -- Chave primária garante idempotência por snapshot + deposito + material
            PRIMARY KEY (snapshot_date, deposito, material)
        );
    """)


def upsert_snapshot(con: sqlite3.Connection, df: pd.DataFrame, snapshot_date: str, source_file: str):
    """
    Upsert (insert/update) do snapshot de estoque no SQLite.

    Entrada:
    - df: DataFrame já normalizado do SAP (load_sap)
      Esperado conter, no mínimo:
        deposito, material, descricao,
        qtd_livre, qtd_qc, qtd_bloq, qtd_total,
        val_livre, val_qc, val_bloq, val_total
    - snapshot_date: string 'YYYY-MM-DD' (data do snapshot que será gravada)
    - source_file: nome do arquivo SAP usado para gerar o snapshot (auditoria)

    Persistência:
    - Tabela: fact_estoque_snapshot
    - Chave primária: (snapshot_date, deposito, material)
    - Em caso de conflito (mesma chave), atualiza os valores (idempotência).
    """

    # Garante que a tabela existe antes de gravar
    ensure_fact_snapshot(con)

    # Timestamp da carga para auditoria
    load_ts = now_ts()

    # ------------------------------------------------------------
    # 1) Monta as linhas (tuplas) para executemany
    # ------------------------------------------------------------
    # Por que criar `rows` manualmente:
    # - evita problemas de tipos do pandas direto no sqlite
    # - melhora performance (executemany)
    # - permite normalização padronizada (deposito/material)
    rows = []
    for _, r in df.iterrows():
        rows.append((
            # Chave do snapshot (com normalização de deposito/material)
            snapshot_date,
            str(r["deposito"]).strip().upper(),
            norm_material_text(r["material"]),

            # Atributo descritivo (não participa da chave)
            r.get("descricao"),

            # Quantidades (None quando vazio/NaN)
            float(r["qtd_livre"]) if pd.notna(r.get("qtd_livre")) else None,
            float(r["qtd_qc"]) if pd.notna(r.get("qtd_qc")) else None,
            float(r["qtd_bloq"]) if pd.notna(r.get("qtd_bloq")) else None,
            float(r["qtd_total"]) if pd.notna(r.get("qtd_total")) else None,

            # Valores (None quando vazio/NaN)
            float(r["val_livre"]) if pd.notna(r.get("val_livre")) else None,
            float(r["val_qc"]) if pd.notna(r.get("val_qc")) else None,
            float(r["val_bloq"]) if pd.notna(r.get("val_bloq")) else None,
            float(r["val_total"]) if pd.notna(r.get("val_total")) else None,

            # Auditoria
            source_file,
            load_ts
        ))

    # sanitize_rows:
    # - garante que as tuplas não tenham tipos inválidos/None inesperados
    rows = sanitize_rows(rows)

    # ------------------------------------------------------------
    # 2) Upsert no SQLite
    # ------------------------------------------------------------
    # ON CONFLICT(snapshot_date, deposito, material):
    # - combina exatamente com a PRIMARY KEY da tabela fact_estoque_snapshot
    #
    # Efeito:
    # - se não existe a linha para aquele material naquele dia -> INSERT
    # - se já existe -> UPDATE (sobrescreve valores com a versão mais recente)
    con.executemany("""
        INSERT INTO fact_estoque_snapshot (
          snapshot_date, deposito, material, descricao,
          qtd_livre, qtd_qc, qtd_bloq, qtd_total,
          val_livre, val_qc, val_bloq, val_total,
          source_file, load_ts
        )
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(snapshot_date, deposito, material) DO UPDATE SET
          descricao=excluded.descricao,
          qtd_livre=excluded.qtd_livre,
          qtd_qc=excluded.qtd_qc,
          qtd_bloq=excluded.qtd_bloq,
          qtd_total=excluded.qtd_total,
          val_livre=excluded.val_livre,
          val_qc=excluded.val_qc,
          val_bloq=excluded.val_bloq,
          val_total=excluded.val_total,
          source_file=excluded.source_file,
          load_ts=excluded.load_ts;
    """, rows)

def ensure_fact_mb51_mov(con: sqlite3.Connection):
    """
    Tabela fato de movimentos MB51 (entradas e saídas).

    Por que é nova:
    - A PK antiga (data_lancamento, deposito, material, ordem) perde movimentos
      (entrada e saída podem compartilhar material+ordem no mesmo dia).
    - Aqui usamos documento+item (quando disponível) para garantir granularidade.

    Chave primária:
      (data_lancamento, deposito, material, doc_material, item_doc_material)

    Observação:
    - Se doc_material/item_doc_material vierem vazios, ainda grava,
      mas a granularidade fica mais fraca. Ideal: MB51 exportar esses campos.
    """
    con.execute("""
        CREATE TABLE IF NOT EXISTS fact_mb51_mov (
            data_lancamento TEXT NOT NULL,
            deposito TEXT NOT NULL,
            material TEXT NOT NULL,

            deb_cred TEXT NOT NULL,       -- 'S' (entrada) / 'H' (saída)
            ordem TEXT,
            tipo_movimento TEXT,

            doc_deposito TEXT,
            doc_material TEXT NOT NULL,
            item_doc_material TEXT NOT NULL,

            descricao_mb51 TEXT,
            quantidade REAL,
            custo_unit REAL,
            valor_estimado REAL,

            source_file TEXT,
            load_ts TEXT,

            PRIMARY KEY (data_lancamento, deposito, material, doc_material, item_doc_material)
        );
    """)


def ensure_fact_mb51_entradas(con: sqlite3.Connection):
    """
    Fato de movimentos MB51 (MAST/depósitos).
    Chave: (data_lancamento, deposito, material, ordem)

    IMPORTANTE (migração):
    - Se a tabela já existir, esta função adiciona novas colunas via ALTER TABLE,
      sem apagar dados.
    """

    # 1) Cria tabela base (caso ainda não exista)
    con.execute("""
        CREATE TABLE IF NOT EXISTS fact_mb51_entradas (
            data_lancamento TEXT NOT NULL,
            deposito TEXT NOT NULL,
            material TEXT NOT NULL,
            ordem TEXT,
            descricao_mb51 TEXT,
            quantidade REAL,
            custo_unit REAL,
            valor_estimado REAL,

            -- NOVAS (para separar entrada/saída e auditoria)
            deb_cred TEXT,          -- 'S' (entrada) / 'H' (saída)
            tipo_movimento TEXT,    -- ex.: 101/102/261 etc (dependendo do layout)
            doc_deposito TEXT,      -- opcional (documento do depósito, se você quiser guardar)
            doc_material TEXT,      -- opcional (doc.material)
            item_doc_material TEXT, -- opcional (item doc.material)

            source_file TEXT,
            load_ts TEXT,

            PRIMARY KEY (data_lancamento, deposito, material, ordem)
        );
    """)

    # 2) Migração leve: adiciona colunas se a tabela já existia sem elas
    cols = {row[1] for row in con.execute("PRAGMA table_info(fact_mb51_entradas);").fetchall()}

    def add_col_if_missing(col_name: str, col_type: str):
        if col_name not in cols:
            con.execute(f"ALTER TABLE fact_mb51_entradas ADD COLUMN {col_name} {col_type};")

    add_col_if_missing("deb_cred", "TEXT")
    add_col_if_missing("tipo_movimento", "TEXT")
    add_col_if_missing("doc_deposito", "TEXT")
    add_col_if_missing("doc_material", "TEXT")
    add_col_if_missing("item_doc_material", "TEXT")


def ensure_mb51_schema(con: sqlite3.Connection):
    """
    Mantém compatibilidade com o nome usado no run().

    Papel:
    - Garantir que a tabela `fact_mb51_entradas` exista e tenha o schema esperado.

    Observação:
    - Nesta V3, a responsabilidade real de criar/ajustar a tabela está em:
      `ensure_fact_mb51_mov(con)`.
    - Esta função é um "alias" para deixar o run() mais legível e evitar divergência
      de nomes entre versões.
    """
    ensure_fact_mb51_mov(con)

def upsert_mb51_mov(con: sqlite3.Connection, df_mb51: pd.DataFrame, source_file: str):
    """
    Upsert de movimentos MB51 na tabela fact_mb51_mov (entradas e saídas).

    - Enriquecimento por custo_unit via dim_material_custo
    - valor_estimado = quantidade * custo_unit
    - PK usa doc_material + item_doc_material para não perder movimentos
    """
    ensure_fact_mb51_mov(con)
    ensure_dim_material_custo(con)

    load_ts = now_ts()

    custo_df = pd.read_sql_query("SELECT material, custo_unit FROM dim_material_custo;", con)
    if not custo_df.empty:
        custo_df["material"] = custo_df["material"].astype("string").apply(norm_material_text)

    tmp = df_mb51.copy()

    tmp["material"] = tmp["material"].astype("string").apply(norm_material_text)
    tmp["deposito"] = tmp["deposito"].astype("string").str.strip().str.upper()
    tmp["ordem"] = tmp["ordem"].fillna("").astype("string").str.strip()
    tmp["deb_cred"] = tmp["deb_cred"].fillna("").astype("string").str.strip().str.upper()

    tmp["tipo_movimento"] = tmp.get("tipo_movimento", "").astype("string").fillna("").str.strip()
    tmp["doc_deposito"] = tmp.get("doc_deposito", "").astype("string").fillna("").str.strip()
    tmp["doc_material"] = tmp.get("doc_material", "").astype("string").fillna("").str.strip()
    tmp["item_doc_material"] = tmp.get("item_doc_material", "").astype("string").fillna("").str.strip()
    tmp["descricao_mb51"] = tmp.get("descricao_mb51", "").astype("string").fillna("").str.strip()

    tmp["data_lancamento"] = pd.to_datetime(tmp["data_lancamento"], errors="coerce", dayfirst=True)
    tmp = tmp.dropna(subset=["data_lancamento"])
    tmp["data_lancamento"] = tmp["data_lancamento"].dt.strftime("%Y-%m-%d")

    # Se doc_material/item_doc_material estiverem vazios, a PK fica fraca.
    # Aqui eu forço a existência para não quebrar o INSERT:
    tmp["doc_material"] = tmp["doc_material"].replace("", pd.NA).fillna("SEM_DOC_MAT")
    tmp["item_doc_material"] = tmp["item_doc_material"].replace("", pd.NA).fillna("SEM_ITEM")

    if not custo_df.empty:
        tmp = tmp.merge(custo_df, on="material", how="left")
    else:
        tmp["custo_unit"] = np.nan

    tmp["quantidade"] = pd.to_numeric(tmp["quantidade"], errors="coerce").fillna(0.0)
    tmp["custo_unit"] = pd.to_numeric(tmp["custo_unit"], errors="coerce").fillna(0.0)
    tmp["valor_estimado"] = tmp["quantidade"] * tmp["custo_unit"]

    rows = []
    for _, r in tmp.iterrows():
        rows.append((
            r.get("data_lancamento") or "",
            str(r.get("deposito") or "").strip().upper(),
            norm_material_text(r.get("material")),

            str(r.get("deb_cred") or "").strip().upper(),
            str(r.get("ordem") or "").strip(),
            str(r.get("tipo_movimento") or "").strip(),

            str(r.get("doc_deposito") or "").strip(),
            str(r.get("doc_material") or "SEM_DOC_MAT").strip(),
            str(r.get("item_doc_material") or "SEM_ITEM").strip(),

            str(r.get("descricao_mb51") or "").strip(),
            float(r.get("quantidade") or 0.0),
            float(r.get("custo_unit") or 0.0),
            float(r.get("valor_estimado") or 0.0),

            source_file,
            load_ts,
        ))

    rows = sanitize_rows(rows)

    con.executemany("""
        INSERT INTO fact_mb51_mov (
            data_lancamento, deposito, material,
            deb_cred, ordem, tipo_movimento,
            doc_deposito, doc_material, item_doc_material,
            descricao_mb51, quantidade, custo_unit, valor_estimado,
            source_file, load_ts
        )
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(data_lancamento, deposito, material, doc_material, item_doc_material) DO UPDATE SET
            deb_cred=excluded.deb_cred,
            ordem=excluded.ordem,
            tipo_movimento=excluded.tipo_movimento,
            doc_deposito=excluded.doc_deposito,
            descricao_mb51=excluded.descricao_mb51,
            quantidade=excluded.quantidade,
            custo_unit=excluded.custo_unit,
            valor_estimado=excluded.valor_estimado,
            source_file=excluded.source_file,
            load_ts=excluded.load_ts;
    """, rows)


def upsert_mb51_entradas(con: sqlite3.Connection, df_mb51: pd.DataFrame, source_file: str):
    """
    Upsert (insert/update) das entradas MB51 no banco.

    Entrada esperada (df_mb51) — *após filtros do run()*:
      - data_lancamento (datetime ou texto convertível)
      - deposito
      - material
      - ordem
      - descricao_mb51
      - quantidade

    Enriquecimento:
      - Faz join com `dim_material_custo` para obter custo_unit por material
      - Calcula valor_estimado = quantidade * custo_unit

    Persistência:
      - Insere em `fact_mb51_entradas`
      - Se já existir a mesma chave (data_lancamento, deposito, material, ordem),
        atualiza os campos via ON CONFLICT DO UPDATE (idempotência por chave).
    """

    # Garante que a tabela fato exista
    ensure_fact_mb51_mov(con)

    # Garante que a dimensão de custo exista (mesmo que esteja vazia)
    ensure_dim_material_custo(con)

    # Timestamp da carga (auditoria)
    load_ts = now_ts()

    # ------------------------------------------------------------
    # 1) Carrega custos do banco (dimensão) para enriquecer o MB51
    # ------------------------------------------------------------
    custo_df = pd.read_sql_query("SELECT material, custo_unit FROM dim_material_custo;", con)

    # Normaliza material na dimensão para casar com o MB51
    if not custo_df.empty:
        custo_df["material"] = custo_df["material"].astype("string").apply(norm_material_text)

    # ------------------------------------------------------------
    # 2) Normalizações no dataframe do MB51 (chaves e tipos)
    # ------------------------------------------------------------
    tmp = df_mb51.copy()

    # material e deposito normalizados para evitar "chaves diferentes" por formatação
    tmp["material"] = tmp["material"].astype("string").apply(norm_material_text)
    tmp["deposito"] = tmp["deposito"].astype("string").str.strip().str.upper()

    # ordem: vazio/NaN vira string vazia (chave do upsert inclui ordem)
    tmp["ordem"] = tmp["ordem"].fillna("").astype("string").str.strip()

    # data_lancamento:
    # - força datetime
    # - depois converte para string "YYYY-MM-DD"
    # Observação: dayfirst=True é comum em relatórios BR
    tmp["data_lancamento"] = pd.to_datetime(tmp["data_lancamento"], errors="coerce", dayfirst=True)
    tmp["data_lancamento"] = tmp["data_lancamento"].dt.strftime("%Y-%m-%d")

    # ------------------------------------------------------------
    # 3) Merge (enriquecimento) com custo_unit
    # ------------------------------------------------------------
    # Se custo_df estiver vazio, custo_unit vira NaN e depois 0.0
    if not custo_df.empty:
        tmp = tmp.merge(custo_df, on="material", how="left")
    else:
        tmp["custo_unit"] = np.nan

    # Converte numéricos de forma defensiva
    tmp["quantidade"] = pd.to_numeric(tmp["quantidade"], errors="coerce").fillna(0.0)
    tmp["custo_unit"] = pd.to_numeric(tmp["custo_unit"], errors="coerce").fillna(0.0)

    # Valor estimado (para análises e agregações no Streamlit)
    tmp["valor_estimado"] = tmp["quantidade"] * tmp["custo_unit"]

    # ------------------------------------------------------------
    # 4) Monta lista de tuplas (rows) para executemany
    # ------------------------------------------------------------
    # Por que montar rows:
    # - executemany é mais performático do que inserir linha a linha via SQL
    # - também evita problemas com tipos do pandas direto no sqlite
    rows = []
    for _, r in tmp.iterrows():
        rows.append((
            r.get("data_lancamento") or "",
            str(r.get("deposito") or "").strip().upper(),
            norm_material_text(r.get("material")),
            str(r.get("ordem") or "").strip(),
            str(r.get("descricao_mb51") or "").strip(),
            float(r.get("quantidade") or 0.0),
            float(r.get("custo_unit") or 0.0),
            float(r.get("valor_estimado") or 0.0),
            source_file,
            load_ts,
            str(r.get("deb_cred") or "").strip(),
            str(r.get("tipo_movimento") or "").strip(),
            str(r.get("doc_deposito") or "").strip(),
            str(r.get("doc_material") or "").strip(),
            str(r.get("item_doc_material") or "").strip(),
))


    # sanitize_rows:
    # - normaliza e garante que não existam None/valores inválidos nas tuplas
    rows = sanitize_rows(rows)

    # ------------------------------------------------------------
    # 5) Upsert no SQLite
    # ------------------------------------------------------------
    # Chave de conflito:
    #   (data_lancamento, deposito, material, ordem)
    #
    # Isso garante idempotência por chave:
    # - reprocessar o mesmo arquivo (ou o mesmo período) não deve duplicar linhas
    #
    # O UPDATE atualiza:
    # - descrição, quantidade, custo e valor, além do source_file e load_ts
    con.executemany("""
        INSERT INTO fact_mb51_entradas (
            data_lancamento, deposito, material, ordem, descricao_mb51,
            quantidade, custo_unit, valor_estimado,
            source_file, load_ts,
            deb_cred, tipo_movimento, doc_deposito, doc_material, item_doc_material
        )
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(data_lancamento, deposito, material, ordem) DO UPDATE SET
            descricao_mb51=excluded.descricao_mb51,
            quantidade=excluded.quantidade,
            custo_unit=excluded.custo_unit,
            valor_estimado=excluded.valor_estimado,
            source_file=excluded.source_file,
            load_ts=excluded.load_ts,
            deb_cred=excluded.deb_cred,
            tipo_movimento=excluded.tipo_movimento,
            doc_deposito=excluded.doc_deposito,
            doc_material=excluded.doc_material,
            item_doc_material=excluded.item_doc_material;
    """, rows)


def ensure_baseline_tables(con: sqlite3.Connection):
    """
    Garante a existência da tabela baseline_mensal.

    baseline_mensal representa o "ponto zero" do mês para cada depósito,
    usado como referência para:
      - cálculo de giro
      - comparação de redução/aumento de estoque
      - KPIs mensais

    Regra estrutural:
    - Uma linha por (mes_ref, deposito)
    - Nunca deve haver mais de um baseline por mês e depósito
    """

    con.execute("""
        CREATE TABLE IF NOT EXISTS baseline_mensal (
            mes_ref TEXT NOT NULL,          -- mês de referência (YYYY-MM)
            deposito TEXT NOT NULL,         -- MAST / MASR / WEPV
            baseline_date TEXT,             -- data do snapshot usado como baseline
            qtd_base_total REAL,            -- quantidade total no baseline
            val_base_total REAL,            -- valor total no baseline
            source_file TEXT,               -- arquivo SAP que originou o baseline
            load_ts TEXT,                   -- timestamp de carga
            PRIMARY KEY (mes_ref, deposito) -- garante 1 baseline por mês/depósito
        );
    """)


def ensure_kpi_table(con: sqlite3.Connection):
    """
    Garante a existência da tabela de KPI mensal resumido.

    Tabela: kpi_mensal_resumo
    Granularidade (chave primária):
      - 1 linha por (mes_ref, deposito)

    Por que essa tabela existe:
    - O Streamlit precisa de um resumo "pronto" por mês e depósito
      para exibir KPIs sem recalcular tudo na interface.
    - Facilita performance e padroniza o resultado do mês.

    Campos (conceito):
    - mes_ref: 'YYYY-MM'
    - deposito: MAST / MASR / WEPV
    - meta_giro: meta definida em config (ex.: 1.2)
    - val_base_total / qtd_base_total:
        * baseline do mês (primeiro dia útil)
    - val_atual_total / qtd_atual_total:
        * valor/quantidade do snapshot "atual" (data d rodada)
    - giro_valor / giro_qtd:
        * indicadores derivados (como você definiu no cálculo do KPI)
    - val_alvo:
        * valor alvo com base na meta (dependente da regra do KPI)
    - falta_baixar:
        * quanto ainda falta reduzir para atingir o alvo (valor)
    - updated_at:
        * timestamp textual de quando o KPI foi recalculado/gravadado
          (ajuda auditoria e debug de "não atualizou no app")
    """
    con.execute("""
        CREATE TABLE IF NOT EXISTS kpi_mensal_resumo (
            mes_ref TEXT NOT NULL,       -- 'YYYY-MM'
            deposito TEXT NOT NULL,      -- MAST/MASR/WEPV

            meta_giro REAL,              -- meta do mês (ex.: 1.2)

            val_base_total REAL,         -- baseline (valor total no primeiro dia útil)
            val_atual_total REAL,        -- valor total no snapshot do dia (d)
            giro_valor REAL,             -- KPI derivado (valor)

            val_alvo REAL,               -- alvo de valor para a meta
            falta_baixar REAL,           -- quanto falta reduzir em valor para atingir o alvo

            qtd_base_total REAL,         -- baseline (quantidade total no primeiro dia útil)
            qtd_atual_total REAL,        -- quantidade total no snapshot do dia (d)
            giro_qtd REAL,               -- KPI derivado (quantidade)

            updated_at TEXT,             -- quando este KPI foi atualizado (auditoria)

            -- Chave primária garante idempotência por mês e depósito:
            -- recalcular o KPI atualiza a mesma linha, não duplica.
            PRIMARY KEY (mes_ref, deposito)
        );
    """)

def ensure_kpi_diario_table(con: sqlite3.Connection):
    """
    KPI diário por depósito, baseado no snapshot + MB51 (MAST).
    Chave: (snapshot_date, deposito)
    """
    con.execute("""
        CREATE TABLE IF NOT EXISTS kpi_diario_deposito (
            snapshot_date TEXT NOT NULL,
            deposito TEXT NOT NULL,

            val_total REAL,
            qtd_total REAL,

            prev_snapshot_date TEXT,
            delta_val REAL,
            delta_qtd REAL,

            entradas_mb51_val REAL,
            entradas_mb51_qtd REAL,

            net_val REAL,
            net_qtd REAL,

            updated_at TEXT,
            PRIMARY KEY (snapshot_date, deposito)
        );
    """)

def ensure_kpi_diario_cols(con: sqlite3.Connection) -> None:
    """
    Migração leve/segura da tabela kpi_diario_deposito.
    Adiciona colunas novas caso o banco já exista (não quebra dados antigos).
    """
    cols = [r[1] for r in con.execute("PRAGMA table_info(kpi_diario_deposito)").fetchall()]

    # Se a tabela ainda não existe, não faz nada aqui.
    # (o ensure_kpi_diario_table(con) deve criar)
    if not cols:
        return

    if "saidas_mb51_val" not in cols:
        con.execute("ALTER TABLE kpi_diario_deposito ADD COLUMN saidas_mb51_val REAL DEFAULT 0;")

    if "saidas_mb51_qtd" not in cols:
        con.execute("ALTER TABLE kpi_diario_deposito ADD COLUMN saidas_mb51_qtd REAL DEFAULT 0;")


def compute_and_store_kpi_diario(con: sqlite3.Connection, deposito: str, d: date):
    """
    Calcula e grava KPI diário do depósito para a data d (snapshot_date = d),
    comparando com o snapshot anterior disponível (qualquer data anterior).

    Regras:
    - snapshot do dia: usa snapshot_date = sd (d em YYYY-MM-DD)
    - snapshot anterior: MAX(snapshot_date) < sd para o mesmo depósito
    - entradas MB51: apenas para MAST, soma fact_mb51_entradas do sd com ordem vazia
    """
    ensure_kpi_diario_table(con)
    ensure_kpi_diario_cols(con)   # <-- ADICIONA ESTA LINHA
    ensure_fact_snapshot(con)
    ensure_fact_mb51_mov(con)


    sd = to_date_str(d)

    # Totais do snapshot do dia (se não existir, não grava)
    row_today = con.execute("""
        SELECT SUM(qtd_total), SUM(val_total)
        FROM fact_estoque_snapshot
        WHERE deposito=? AND snapshot_date=?;
    """, (deposito, sd)).fetchone()

    if not row_today:
        return None

    qtd_today, val_today = row_today
    if qtd_today is None and val_today is None:
        return None

    # Snapshot anterior disponível
    prev_date = con.execute("""
        SELECT MAX(snapshot_date)
        FROM fact_estoque_snapshot
        WHERE deposito=? AND snapshot_date < ?;
    """, (deposito, sd)).fetchone()[0]

    qtd_prev = None
    val_prev = None
    if prev_date:
        row_prev = con.execute("""
            SELECT SUM(qtd_total), SUM(val_total)
            FROM fact_estoque_snapshot
            WHERE deposito=? AND snapshot_date=?;
        """, (deposito, prev_date)).fetchone()
        if row_prev:
            qtd_prev, val_prev = row_prev

    # Delta vs anterior (se não existir anterior, delta = None)
    delta_val = (val_today - val_prev) if (val_today is not None and val_prev is not None) else None
    delta_qtd = (qtd_today - qtd_prev) if (qtd_today is not None and qtd_prev is not None) else None


    # ------------------------------------------------------------
    # MB51 (MAST): regra de negócio para eficiência
    # Entradas válidas: deb_cred='S' e ordem vazia (sem OP)
    # Saídas válidas:   deb_cred='H' e ordem preenchida (com OP)
    # net_val/net_qtd:  consumo líquido = saídas - entradas
    # ------------------------------------------------------------
    ent_in_val = 0.0
    ent_in_qtd = 0.0
    sai_val = 0.0
    sai_qtd = 0.0

    if str(deposito).strip().upper() == "MAST":
        row_in = con.execute("""
            SELECT
                COALESCE(SUM(valor_estimado), 0),
                COALESCE(SUM(quantidade), 0)
            FROM fact_mb51_mov
            WHERE deposito = ?
            AND data_lancamento = ?
            AND COALESCE(deb_cred,'') = 'S'
            AND TRIM(COALESCE(ordem,'')) = '';
        """, (str(deposito).strip().upper(), sd)).fetchone()

        ent_in_val = float(row_in[0] or 0.0)
        ent_in_qtd = float(row_in[1] or 0.0)

        row_out = con.execute("""
            SELECT
                COALESCE(SUM(valor_estimado), 0),
                COALESCE(SUM(quantidade), 0)
            FROM fact_mb51_mov
            WHERE deposito = ?
            AND data_lancamento = ?
            AND COALESCE(deb_cred,'') = 'H'
            AND TRIM(COALESCE(ordem,'')) <> '';
        """, (str(deposito).strip().upper(), sd)).fetchone()

        sai_val = float(row_out[0] or 0.0)
        sai_qtd = float(row_out[1] or 0.0)

    # Entradas (para manter o nome atual do campo)
    ent_val = ent_in_val
    ent_qtd = ent_in_qtd

    # ------------------------------------------------------------
    # KPI principal de eficiência (MB51):
    # consumo líquido = Saídas - Entradas
    # ------------------------------------------------------------
    net_val = None
    net_qtd = None

    if str(deposito).strip().upper() == "MAST":
        net_val = sai_val - ent_in_val
        net_qtd = sai_qtd - ent_in_qtd

    # ------------------------------------------------------------
    # Métrica auxiliar (auditoria) baseada no snapshot:
    # delta = entradas - saídas  =>  saídas_est = entradas - delta
    # ------------------------------------------------------------
    consumo_estimado_val = None
    consumo_estimado_qtd = None

    if str(deposito).strip().upper() == "MAST":
        if delta_val is not None:
            consumo_estimado_val = ent_in_val - delta_val
        if delta_qtd is not None:
            consumo_estimado_qtd = ent_in_qtd - delta_qtd

    # Entradas (mantém compatibilidade com nomes existentes no INSERT)
    ent_val = ent_in_val
    ent_qtd = ent_in_qtd

    # Arredondamentos finais (opcional, mas ajuda)
    ent_val = round(float(ent_val or 0.0), 2)
    ent_qtd = round(float(ent_qtd or 0.0), 3)
    sai_val = round(float(sai_val or 0.0), 2)
    sai_qtd = round(float(sai_qtd or 0.0), 3)

    if net_val is not None:
        net_val = round(float(net_val), 2)
    if net_qtd is not None:
        net_qtd = round(float(net_qtd), 3)

    if consumo_estimado_val is not None:
        consumo_estimado_val = round(float(consumo_estimado_val), 2)
    if consumo_estimado_qtd is not None:
        consumo_estimado_qtd = round(float(consumo_estimado_qtd), 3)

    sai_val = round(float(sai_val or 0.0), 2)
    sai_qtd = round(float(sai_qtd or 0.0), 3)


    con.execute("""
    INSERT INTO kpi_diario_deposito (
        snapshot_date, deposito,
        val_total, qtd_total,
        prev_snapshot_date, delta_val, delta_qtd,

        entradas_mb51_val, entradas_mb51_qtd,
        saidas_mb51_val, saidas_mb51_qtd,

        net_val, net_qtd,
        updated_at
    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    ON CONFLICT(snapshot_date, deposito) DO UPDATE SET
        val_total=excluded.val_total,
        qtd_total=excluded.qtd_total,
        prev_snapshot_date=excluded.prev_snapshot_date,
        delta_val=excluded.delta_val,
        delta_qtd=excluded.delta_qtd,

        entradas_mb51_val=excluded.entradas_mb51_val,
        entradas_mb51_qtd=excluded.entradas_mb51_qtd,
        saidas_mb51_val=excluded.saidas_mb51_val,
        saidas_mb51_qtd=excluded.saidas_mb51_qtd,

        net_val=excluded.net_val,
        net_qtd=excluded.net_qtd,
        updated_at=excluded.updated_at;
    """, (
    sd, str(deposito).strip().upper(),
    val_today, qtd_today,
    prev_date, delta_val, delta_qtd,

    ent_val, ent_qtd,
    sai_val, sai_qtd,

    net_val, net_qtd,
    now_ts()
    ))

    return True


def backfill_kpi_diario_mes(con: sqlite3.Connection, deposito: str, d: date):
    """
    Gera/atualiza o KPI diário do depósito para TODAS as datas de snapshot
    do mês de referência até a data d.

    Por que existe:
    - o ETL roda em um dia (d), mas o Streamlit precisa da série diária do mês.
    - então recalculamos o diário para cada snapshot_date disponível no mês.

    Segurança:
    - usa UPSERT em kpi_diario_deposito, então não duplica.
    """

    ensure_kpi_diario_table(con)

    mref = mes_ref(d)
    sd = to_date_str(d)

    # Lista todas as datas de snapshot do mês para o depósito, até a data d
    dates = con.execute("""
        SELECT snapshot_date
        FROM fact_estoque_snapshot
        WHERE deposito=? AND snapshot_date LIKE ? AND snapshot_date <= ?
        GROUP BY snapshot_date
        ORDER BY snapshot_date;
    """, (deposito, f"{mref}-%", sd)).fetchall()

    # fetchall -> lista de tuplas (('2026-01-02',), ...)
    for (snap_date,) in dates:
        try:
            snap_d = datetime.strptime(snap_date, "%Y-%m-%d").date()
            compute_and_store_kpi_diario(con, deposito, snap_d)
        except Exception as e:
            # Não quebra o ETL inteiro por causa de 1 dia; apenas registra
            print_step(f"KPI_DIARIO: erro em {deposito} {snap_date}: {type(e).__name__}: {e}")


def ensure_baseline(con: sqlite3.Connection, deposito: str, d: date, source_file: str) -> bool:
    """
    Garante a existência do baseline do mês (baseline_mensal) para um depósito.

    Conceito:
    - Baseline = "ponto de partida" do mês para comparar evolução e calcular KPI.
    - Regra do projeto: baseline deve representar o mês no primeiro dia útil.

    Estratégia usada (V3):
    1) Se já existe baseline para (mes_ref, deposito), não faz nada (retorna False).
    2) Determina o primeiro dia útil do mês (first_business_day(d)).
    3) Se estamos exatamente no primeiro dia útil:
       - baseline_date = data de hoje (d)
    4) Se não estamos no primeiro dia útil:
       - baseline_date = menor snapshot_date disponível no mês para o depósito
         (primeiro snapshot gravado do mês)
       - Se não existir snapshot no mês, não cria baseline.

    Nota importante:
    - Esta função depende de fact_estoque_snapshot já estar populada.
    - Ela é chamada no run() depois do upsert_snapshot, então normalmente existe snapshot.

    Retorno:
    - True  => baseline criado agora
    - False => baseline já existia ou não foi possível criar (sem dados)
    """

    # Garante que tabelas de baseline existem
    ensure_baseline_tables(con)

    # Mês referência (YYYY-MM)
    mref = mes_ref(d)

    # ------------------------------------------------------------
    # 1) Se baseline já existe para o mês/depósito, não recria
    # ------------------------------------------------------------
    cur = con.execute(
        "SELECT 1 FROM baseline_mensal WHERE mes_ref=? AND deposito=? LIMIT 1",
        (mref, deposito)
    )
    if cur.fetchone():
        return False

    # ------------------------------------------------------------
    # 2) Define baseline_date
    # ------------------------------------------------------------
    # first_business_day(d) deve retornar o primeiro dia útil DO MÊS de d
    bday = first_business_day(d)

    # Se hoje é o primeiro dia útil do mês, baseline é hoje
    if d == bday:
        baseline_date = to_date_str(d)

    # Caso contrário, usamos o primeiro snapshot disponível no mês (MIN(snapshot_date))
    else:
        baseline_date = con.execute("""
            SELECT MIN(snapshot_date)
            FROM fact_estoque_snapshot
            WHERE deposito=? AND snapshot_date LIKE ?;
        """, (deposito, f"{mref}-%")).fetchone()[0]

        # Se não há snapshot no mês, não dá para criar baseline
        if not baseline_date:
            return False

    # ------------------------------------------------------------
    # 3) Calcula totais do baseline (qtd_total e val_total)
    # ------------------------------------------------------------
    row = con.execute("""
        SELECT SUM(qtd_total) AS qtd_total, SUM(val_total) AS val_total
        FROM fact_estoque_snapshot
        WHERE snapshot_date=? AND deposito=?;
    """, (baseline_date, deposito)).fetchone()

    qtd_base = row[0] if row else None
    val_base = row[1] if row else None

    # ------------------------------------------------------------
    # 4) Insere baseline_mensal
    # ------------------------------------------------------------
    con.execute("""
        INSERT INTO baseline_mensal (
            mes_ref, deposito, baseline_date, qtd_base_total, val_base_total, source_file, load_ts
        ) VALUES (?,?,?,?,?,?,?)
    """, (
        mref, deposito, baseline_date,
        qtd_base, val_base,
        source_file,
        now_ts()
    ))

    # Log para auditoria (ótimo para explicar o comportamento para supervisão)
    print_step(f"BASELINE criado para {deposito} em {mref}: baseline_date={baseline_date}")
    return True


def compute_and_store_kpi(con: sqlite3.Connection, deposito: str, d: date, meta_giro: float):
    """
    Calcula e grava o KPI mensal consolidado na tabela kpi_mensal_resumo.

    Entradas:
    - deposito: 'MAST' / 'MASR' / 'WEPV'
    - d: data de referência da execução (date)
    - meta_giro: meta configurada (ex.: 1.2)

    Lógica de alto nível:
    1) Determina o mês de referência (mes_ref = 'YYYY-MM')
    2) Busca o baseline do mês (baseline_mensal) para o depósito
       - baseline = primeiro dia útil do mês (garantido pela ensure_baseline)
    3) Descobre o último snapshot existente no mês até a data d
       - isso protege o KPI quando não houve snapshot no dia exato d
    4) Soma qtd_total e val_total do último snapshot encontrado
    5) Calcula métricas:
       - giro_valor = val_base / val_atual
       - giro_qtd   = qtd_base / qtd_atual
       - val_alvo   = val_base / meta_giro
       - falta_baixar = val_atual - val_alvo
    6) Faz UPSERT em kpi_mensal_resumo (1 linha por mês e depósito)
       - atualiza updated_at para auditoria

    Retorno:
    - True quando gravou KPI com sucesso
    - None quando não há baseline ou não há snapshot (não grava nada)
    """

    # Garante que tabelas existam
    ensure_baseline_tables(con)
    ensure_kpi_table(con)

    # ------------------------------------------------------------
    # 1) Mês de referência do KPI (YYYY-MM)
    # ------------------------------------------------------------
    mref = mes_ref(d)

    # ------------------------------------------------------------
    # 2) Busca baseline do mês (quantidade e valor)
    # ------------------------------------------------------------
    # baseline_mensal deve ter 1 linha por (mes_ref, deposito)
    b = con.execute(
        "SELECT qtd_base_total, val_base_total FROM baseline_mensal WHERE mes_ref=? AND deposito=?",
        (mref, deposito)
    ).fetchone()

    # Se ainda não existe baseline, não há como calcular KPI do mês
    if not b:
        return None

    qtd_base, val_base = b

    # ------------------------------------------------------------
    # 3) Define "data limite" e encontra o último snapshot disponível
    # ------------------------------------------------------------
    # sd é a data 'd' em formato 'YYYY-MM-DD'
    sd = to_date_str(d)

    # Estratégia:
    # - procura o MAX(snapshot_date) no mês (mref-%)
    # - garante que snapshot_date <= sd (não usa datas "no futuro")
    #
    # Isso resolve o caso:
    # - rodou ETL em um dia sem snapshot, mas existe snapshot anterior no mês
    last_date = con.execute("""
        SELECT MAX(snapshot_date) FROM fact_estoque_snapshot
        WHERE deposito=? AND snapshot_date LIKE ? AND snapshot_date<=?;
    """, (deposito, f"{mref}-%", sd)).fetchone()[0]

    # Se não existe snapshot no mês até a data d, não calcula KPI
    if not last_date:
        return None

    # ------------------------------------------------------------
    # 4) Totais do snapshot (soma de qtd_total e val_total)
    # ------------------------------------------------------------
    totals = con.execute("""
        SELECT SUM(qtd_total), SUM(val_total)
        FROM fact_estoque_snapshot
        WHERE deposito=? AND snapshot_date=?;
    """, (deposito, last_date)).fetchone()

    qtd_atual, val_atual = totals

    # ------------------------------------------------------------
    # 5) Cálculos do KPI (defensivo contra None/0)
    # ------------------------------------------------------------
    # giro_valor:
    # - quanto maior, melhor (base / atual)
    giro_valor = (val_base / val_atual) if (val_base is not None and val_atual not in (None, 0)) else None

    # giro_qtd:
    # - mesmo raciocínio para quantidade
    giro_qtd = (qtd_base / qtd_atual) if (qtd_base is not None and qtd_atual not in (None, 0)) else None

    # val_alvo:
    # - valor alvo para atingir a meta de giro
    # Ex.: meta 1.2 => alvo = base/1.2 (redução esperada em relação ao baseline)
    val_alvo = (val_base / meta_giro) if (val_base is not None and meta_giro not in (None, 0)) else None

    # falta_baixar:
    # - se val_atual > val_alvo => falta baixar (positivo)
    # - se val_atual < val_alvo => "já bateu meta" (negativo)
    falta_baixar = (val_atual - val_alvo) if (val_atual is not None and val_alvo is not None) else None

    # ------------------------------------------------------------
    # 6) UPSERT no resumo mensal
    # ------------------------------------------------------------
    # ON CONFLICT(mes_ref, deposito):
    # - garante 1 linha por mês e depósito
    # - recalcular KPI atualiza os campos e o updated_at (sem duplicar)
    con.execute("""
      INSERT INTO kpi_mensal_resumo (
        mes_ref, deposito, meta_giro,
        val_base_total, val_atual_total, giro_valor, val_alvo, falta_baixar,
        qtd_base_total, qtd_atual_total, giro_qtd,
        updated_at
      ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
      ON CONFLICT(mes_ref, deposito) DO UPDATE SET
        meta_giro=excluded.meta_giro,
        val_base_total=excluded.val_base_total,
        val_atual_total=excluded.val_atual_total,
        giro_valor=excluded.giro_valor,
        val_alvo=excluded.val_alvo,
        falta_baixar=excluded.falta_baixar,
        qtd_base_total=excluded.qtd_base_total,
        qtd_atual_total=excluded.qtd_atual_total,
        giro_qtd=excluded.giro_qtd,
        updated_at=excluded.updated_at;
    """, (
        mref, deposito, meta_giro,
        val_base, val_atual, giro_valor, val_alvo, falta_baixar,
        qtd_base, qtd_atual, giro_qtd,
        now_ts()
    ))

    return True


def aplicar_blacklist(df: pd.DataFrame) -> pd.Series:
    """
    Aplica a blacklist operacional do MAST.

    O campo processamento agora vem normalizado pelo ASSIST, sem os
    prefixos numéricos usados anteriormente pela base 4MDG.
    """
    item = pd.to_numeric(df["material"], errors="coerce")
    desc = (
        df["descricao"]
        .astype("string")
        .fillna("")
        .str.strip()
    )
    proc = (
        df.get(
            "processamento",
            pd.Series([pd.NA] * len(df), index=df.index),
        )
        .astype("string")
        .fillna("")
        .str.strip()
        .str.upper()
    )

    cond_1 = proc.eq("DESCARTE")

    cond_2 = proc.eq("TESTE CONSERTO")
    cond_3 = proc.eq("DESCARTE DEFEITUOSO")
    cond_4 = proc.eq("TRATATIVA DIFERENCIADA")
    cond_5 = proc.eq("CANIBALIZAÇÃO")
    cond_6 = proc.eq("TESTE")
    cond_7 = proc.eq("REPARO")

    cond_8 = item.isin([1400138, 1085236])

    cond_faixa_187 = item.between(
        1870000,
        1879999,
        inclusive="both",
    )
    cond_desc_inversor = desc.str.contains(
        "inversor",
        case=False,
        na=False,
    )
    cond_desc_microinv = desc.str.contains(
        "microinversor",
        case=False,
        na=False,
    )
    cond_9 = cond_faixa_187 & (
        cond_desc_inversor | cond_desc_microinv
    )

    cond_10 = item.lt(1880000)

    condicoes = [
        cond_1,
        cond_2,
        cond_3,
        cond_4,
        cond_5,
        cond_6,
        cond_7,
        cond_8,
        cond_9,
        cond_10,
    ]
    escolhas = [
        "SIM",
        "NÃO",
        "NÃO",
        "NÃO",
        "NÃO",
        "NÃO",
        "NÃO",
        "NÃO",
        "NÃO",
        "SIM",
    ]

    return pd.Series(
        np.select(
            condicoes,
            escolhas,
            default="NÃO",
        ),
        index=df.index,
        dtype="string",
    )


# ============================================================
# OUTPUT (Excel)
# ============================================================

def build_output(con: sqlite3.Connection, deposito: str, d: date, out_dir: Path, cfg: Dict[str, Any]) -> Path:
    ensure_kpi_table(con)
    ensure_mb51_schema(con)  # garante a tabela existir antes de SELECT no build_output

    mref = mes_ref(d)

    # -----------------------------
    # 1) RESUMO (KPI mensal)
    # -----------------------------
    k = con.execute("""
      SELECT mes_ref, deposito, meta_giro, val_base_total, val_atual_total, giro_valor, val_alvo, falta_baixar,
             qtd_base_total, qtd_atual_total, giro_qtd, updated_at
      FROM kpi_mensal_resumo
      WHERE mes_ref=? AND deposito=?;
    """, (mref, deposito)).fetchone()

    if not k:
        print_step(
            f"ATENÇÃO: KPI não disponível para {deposito} em {mref} "
            "(baseline pode não existir ainda). Gerando OUTPUT sem KPI."
        )
        resumo = pd.DataFrame([{
            "mes_ref": mref,
            "deposito": deposito,
            "meta_giro": None,
            "val_base_total": None,
            "val_atual_total": None,
            "giro_valor": None,
            "val_alvo": None,
            "falta_baixar": None,
            "qtd_base_total": None,
            "qtd_atual_total": None,
            "giro_qtd": None,
            "updated_at": now_ts(),
        }])
    else:
        resumo = pd.DataFrame([{
            "mes_ref": k[0],
            "deposito": k[1],
            "meta_giro": k[2],
            "val_base_total": k[3],
            "val_atual_total": k[4],
            "giro_valor": k[5],
            "val_alvo": k[6],
            "falta_baixar": k[7],
            "qtd_base_total": k[8],
            "qtd_atual_total": k[9],
            "giro_qtd": k[10],
            "updated_at": k[11],
        }])

    # Indicadores extras (MAST)
    if deposito == "MAST":
        val_base = resumo.loc[0, "val_base_total"]
        val_atual = resumo.loc[0, "val_atual_total"]
        qtd_base = resumo.loc[0, "qtd_base_total"]
        qtd_atual = resumo.loc[0, "qtd_atual_total"]

        red_val = safe_ratio(
            (val_base - val_atual) if (val_base is not None and val_atual is not None) else None,
            val_base
        )
        red_qtd = safe_ratio(
            (qtd_base - qtd_atual) if (qtd_base is not None and qtd_atual is not None) else None,
            qtd_base
        )
        indice = (0.70 * red_val + 0.30 * red_qtd) if (red_val is not None and red_qtd is not None) else None

        resumo.loc[0, "red_val_pct"] = red_val
        resumo.loc[0, "red_qtd_pct"] = red_qtd
        resumo.loc[0, "indice_eficiencia_70_30"] = indice

        entradas_mes = con.execute("""
            SELECT
                COALESCE(SUM(quantidade), 0) AS entradas_qtd,
                COALESCE(SUM(valor_estimado), 0) AS entradas_val
            FROM fact_mb51_mov
            WHERE deposito='MAST'
              AND data_lancamento LIKE ?
              AND COALESCE(TRIM(ordem), '') = '';
        """, (f"{mref}-%",)).fetchone()

        resumo.loc[0, "entradas_mast_qtd_mes"] = float(entradas_mes[0] or 0)
        resumo.loc[0, "entradas_mast_val_mes"] = float(entradas_mes[1] or 0)

    # -----------------------------
    # 2) EVOLUÇÃO DO MÊS
    # -----------------------------
    evo = pd.read_sql_query("""
      SELECT snapshot_date, SUM(val_total) AS val_total, SUM(qtd_total) AS qtd_total
      FROM fact_estoque_snapshot
      WHERE deposito=? AND snapshot_date LIKE ?
      GROUP BY snapshot_date
      ORDER BY snapshot_date;
    """, con, params=(deposito, f"{mref}-%"))

    if evo.empty:
        raise RuntimeError("Sem snapshots para o mês atual (evolução vazia).")

    last_date = str(evo["snapshot_date"].max())
    print_step(f"OUTPUT: last_date={last_date} deposito={deposito} mes_ref={mref}")

    out_cfg = cfg.get("output", {}) if isinstance(cfg, dict) else {}
    gerar_base = bool(out_cfg.get("gerar_base_consumo", True))
    gerar_top = bool(out_cfg.get("gerar_top", True))
    top_n = int(out_cfg.get("top_n", 200))
    top_order_by = str(out_cfg.get("top_order_by", "val_total")).strip().lower()

    if top_n < 1:
        top_n = 200
    if top_order_by not in ("val_total", "qtd_total", "valor_total_custo"):
        top_order_by = "val_total"

    # -----------------------------
    # 3) BASE_CONSUMO (último snapshot)
    # -----------------------------
    base_full = pd.read_sql_query("""
      SELECT material, descricao, val_total, qtd_total
      FROM fact_estoque_snapshot
      WHERE deposito=? AND snapshot_date=?;
    """, con, params=(deposito, last_date))

    if base_full.empty:
        raise RuntimeError(f"Sem dados no último snapshot ({last_date}) para depósito {deposito}.")

    base_full["material"] = base_full["material"].astype("string").apply(norm_material_text)

    # Dimensão consolidada de materiais.
    # Fonte comercial: CSV oficial do BI.
    # Hierarquia: dim_segmento.
    # Processamento: ASSIST.
    dim_material = pd.read_sql_query(
        """
        SELECT
            material,
            bu,
            diretoria,
            segmento,
            centro_custo,
            centro_lucro,
            codigo_centro_lucro,
            codigo_centro_sap,
            processamento
        FROM dim_material;
        """,
        con,
    )

    if not dim_material.empty:
        dim_material["material"] = (
            dim_material["material"]
            .astype("string")
            .apply(norm_material_text)
        )

    base_full = base_full.merge(
        dim_material,
        on="material",
        how="left",
    )

    # Compatibilidade temporária com os outputs e páginas atuais:
    # grupo      <- BU
    # familia    <- Diretoria
    # subfamilia <- Segmento
    # uni_neg    <- BU
    base_full["grupo"] = (
        base_full.get("bu", pd.Series(pd.NA, index=base_full.index))
        .fillna("SEM_BU")
    )
    base_full["familia"] = (
        base_full.get("diretoria", pd.Series(pd.NA, index=base_full.index))
        .fillna("SEM_DIRETORIA")
    )
    base_full["subfamilia"] = (
        base_full.get("segmento", pd.Series(pd.NA, index=base_full.index))
        .fillna("SEM_SEGMENTO")
    )
    base_full["uni_neg"] = base_full["grupo"]

    # Campo legado ainda exigido pelo contrato de saída.
    # A nova arquitetura ainda não possui fonte oficial para esse atributo.
    base_full["saida_manufatura"] = pd.NA

    # -----------------------------
    # 4) Dimensão de custo + ALERTAS + COBERTURA
    # -----------------------------
    ensure_dim_material_custo(con)

    # Config do custo (limite de idade em dias)
    # - você pode definir no config.json:  "custo_max_age_days": 30
    custo_max_age_days = int(cfg.get("custo_max_age_days", 30))

    # Verifica se o arquivo configurado existe (serve para seu PC e o servidor)
    custo_file_cfg = str(cfg.get("custo_file", "") or "").strip()
    custo_file_exists = Path(custo_file_cfg).exists() if custo_file_cfg else False

    # Puxa auditoria da dimensão de custo
    # (pega “o mais recente” disponível na dimensão)
    custo_meta = con.execute("""
        SELECT
            MAX(COALESCE(source_file,'')) AS source_file,
            MAX(COALESCE(source_last_modified,'')) AS source_last_modified,
            MAX(COALESCE(load_ts,'')) AS load_ts
        FROM dim_material_custo;
    """).fetchone()

    custo_source_file = (custo_meta[0] or "").strip()
    custo_source_last_modified = (custo_meta[1] or "").strip()
    custo_dim_load_ts = (custo_meta[2] or "").strip()

    # Carrega custos para merge
    custo_df = pd.read_sql_query("SELECT material, custo_unit FROM dim_material_custo", con)
    if not custo_df.empty:
        custo_df["material"] = custo_df["material"].astype("string").apply(norm_material_text)
    base_full = base_full.merge(custo_df, on="material", how="left")

    # custo_unit pode ficar NaN -> mantém NaN para cobertura, e usa 0 só no cálculo
    qtd_num = pd.to_numeric(base_full["qtd_total"], errors="coerce")
    val_num = pd.to_numeric(base_full["val_total"], errors="coerce")
    custo_num = pd.to_numeric(base_full.get("custo_unit", np.nan), errors="coerce")

    has_custo = custo_num.notna() & (custo_num > 0)

    total_materiais = int(len(base_full))
    materiais_com_custo = int(has_custo.sum())
    cobertura_custo_pct_materiais = (materiais_com_custo / total_materiais) if total_materiais > 0 else None

    total_val = float(val_num.fillna(0).sum())
    val_com_custo = float(val_num.where(has_custo, 0).fillna(0).sum())
    cobertura_custo_pct_val = (val_com_custo / total_val) if total_val > 0 else None

    total_qtd = float(qtd_num.fillna(0).sum())
    qtd_com_custo = float(qtd_num.where(has_custo, 0).fillna(0).sum())
    cobertura_custo_pct_qtd = (qtd_com_custo / total_qtd) if total_qtd > 0 else None

    # Idade do custo (se conseguir parsear)
    custo_age_days = None
    custo_is_stale = False
    if custo_source_last_modified:
        try:
            # source_last_modified é string "YYYY-MM-DD HH:MM:SS" no seu ETL
            dt_last = datetime.strptime(custo_source_last_modified, "%Y-%m-%d %H:%M:%S")
            custo_age_days = (datetime.now() - dt_last).days
            custo_is_stale = (custo_age_days is not None and custo_age_days > custo_max_age_days)
        except Exception:
            # Se não parsear, não chuta; apenas registra que não deu.
            custo_age_days = None
            custo_is_stale = False

    # Decide alerta de custo (ordem importa)
    if not custo_file_cfg:
        alerta_custo = "ALERTA: custo_file NÃO configurado no config.json"
    elif not custo_file_exists:
        alerta_custo = "ALERTA: arquivo de custo NÃO encontrado no caminho do config"
    elif custo_df.empty:
        alerta_custo = "ALERTA: dim_material_custo vazia (sem custos carregados)"
    elif custo_is_stale:
        alerta_custo = f"ALERTA: custo desatualizado (>{custo_max_age_days} dias)"
    else:
        alerta_custo = "OK"

    # Grava auditoria e cobertura no RESUMO (fica visível no Excel)
    resumo.loc[0, "alerta_custo"] = alerta_custo
    resumo.loc[0, "custo_file_cfg"] = custo_file_cfg or None
    resumo.loc[0, "custo_file_exists"] = bool(custo_file_exists)
    resumo.loc[0, "custo_source_file"] = custo_source_file or None
    resumo.loc[0, "custo_source_last_modified"] = custo_source_last_modified or None
    resumo.loc[0, "custo_dim_load_ts"] = custo_dim_load_ts or None
    resumo.loc[0, "custo_max_age_days"] = custo_max_age_days
    resumo.loc[0, "custo_age_days"] = custo_age_days
    resumo.loc[0, "cobertura_custo_pct_materiais"] = cobertura_custo_pct_materiais
    resumo.loc[0, "cobertura_custo_pct_val"] = cobertura_custo_pct_val
    resumo.loc[0, "cobertura_custo_pct_qtd"] = cobertura_custo_pct_qtd

    print_step(
    f"CUSTO({deposito}): alerta={alerta_custo} | "
    f"file_exists={custo_file_exists} | "
    f"cov_val={resumo.loc[0,'cobertura_custo_pct_val']} | "
    f"cov_qtd={resumo.loc[0,'cobertura_custo_pct_qtd']} | "
    f"cov_mat={resumo.loc[0,'cobertura_custo_pct_materiais']}"
)

    # Cálculo do custo total (para análises e ordenação top)
    # - aqui sim, usa 0 quando não tem custo
    base_full["valor_total_custo"] = (
        pd.to_numeric(base_full["qtd_total"], errors="coerce").fillna(0) *
        pd.to_numeric(base_full.get("custo_unit", 0), errors="coerce").fillna(0)
    )

    # Campo legado (mantém)
    if "cod_ref_item_obsoleto" not in base_full.columns:
        base_full["cod_ref_item_obsoleto"] = None

    # Blacklist
    base_full["blacklist"] = aplicar_blacklist(base_full)

    # -----------------------------
    # 5) PRIORIDADE (MAST)
    # -----------------------------
    prioridade_df = None
    if deposito == "MAST":
        prioridade_df = base_full.copy()

        qtd_num2 = pd.to_numeric(prioridade_df["qtd_total"], errors="coerce").fillna(0)
        val_custo_num = pd.to_numeric(prioridade_df["valor_total_custo"], errors="coerce").fillna(0)

        total_qtd_dep = float(qtd_num2.sum())
        total_val_dep = float(val_custo_num.sum())

        prioridade_df["pct_qtd_deposito"] = qtd_num2.apply(lambda x: safe_ratio(x, total_qtd_dep))
        prioridade_df["pct_valor_deposito"] = val_custo_num.apply(lambda x: safe_ratio(x, total_val_dep))

        prioridade_df["score_peso_70_30"] = (
            0.70 * prioridade_df["pct_valor_deposito"].fillna(0) +
            0.30 * prioridade_df["pct_qtd_deposito"].fillna(0)
        )
        prioridade_df["score_contra_vies"] = (
            0.70 * prioridade_df["pct_qtd_deposito"].fillna(0) +
            0.30 * prioridade_df["pct_valor_deposito"].fillna(0)
        )

        prioridade_df = prioridade_df[
            prioridade_df["blacklist"].fillna("NÃO").astype("string").str.strip().str.upper().eq("NÃO")
        ].copy()

        ABC_FIELD = "score_contra_vies"
        ABC_A = 0.80
        ABC_B = 0.95

        prioridade_df["_score_abc"] = pd.to_numeric(prioridade_df[ABC_FIELD], errors="coerce").fillna(0).astype(float)

        total_score = float(prioridade_df["_score_abc"].sum())
        if total_score > 0:
            prioridade_df = prioridade_df.sort_values(by="_score_abc", ascending=False)
            prioridade_df["abc_acum_pct"] = prioridade_df["_score_abc"].cumsum() / total_score

            def class_abc(x):
                if x <= ABC_A:
                    return "A"
                if x <= ABC_B:
                    return "B"
                return "C"

            prioridade_df["ABC_contra_vies"] = prioridade_df["abc_acum_pct"].apply(class_abc)
        else:
            prioridade_df["abc_acum_pct"] = 0
            prioridade_df["ABC_contra_vies"] = "C"

        prioridade_df = prioridade_df.drop(columns=["_score_abc"])
        prioridade_df = prioridade_df.sort_values(by="score_contra_vies", ascending=False)

    # -----------------------------
    # 6) TOP
    # -----------------------------
    top_df = base_full.copy()
    if top_order_by not in top_df.columns:
        top_order_by = "val_total"
    top_df = top_df.sort_values(by=top_order_by, ascending=False).head(top_n)

    # -----------------------------
    # 7) ENTRADAS (MAST)
    # -----------------------------
    entradas_df = pd.DataFrame()
    entradas_dia_df = pd.DataFrame()
    if deposito == "MAST":
        entradas_df = pd.read_sql_query("""
            SELECT
                data_lancamento,
                material,
                descricao_mb51,
                quantidade,
                custo_unit,
                valor_estimado,
                ordem
            FROM fact_mb51_mov
            WHERE deposito='MAST'
              AND data_lancamento LIKE ?
              AND COALESCE(TRIM(ordem), '') = ''
            ORDER BY data_lancamento DESC, valor_estimado DESC;
        """, con, params=(f"{mref}-%",))

        entradas_dia_df = pd.read_sql_query("""
            SELECT
                data_lancamento,
                SUM(quantidade) AS entradas_qtd,
                SUM(valor_estimado) AS entradas_valor_estimado
            FROM fact_mb51_mov
            WHERE deposito='MAST'
              AND data_lancamento LIKE ?
              AND COALESCE(TRIM(ordem), '') = ''
            GROUP BY data_lancamento
            ORDER BY data_lancamento;
        """, con, params=(f"{mref}-%",))

    # -----------------------------
    # 8) Contrato de colunas
    # -----------------------------
    cols_contrato = [
        "material", "descricao", "val_total", "qtd_total",
        "grupo", "processamento", "saida_manufatura",
        "custo_unit", "valor_total_custo",
        "uni_neg", "familia", "subfamilia",
        "cod_ref_item_obsoleto",
        "blacklist"
    ]
    for c in cols_contrato:
        if c not in base_full.columns:
            base_full[c] = None
        if c not in top_df.columns:
            top_df[c] = None

    base_full = base_full[cols_contrato]
    top_df = top_df[cols_contrato]

    # -----------------------------
    # 9) Escrita do Excel + fallback
    # -----------------------------
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{deposito}_resumo_{mref}.xlsx"
    temp_path = out_dir / f"{deposito}_resumo_{mref}__TEMP.xlsx"

    resumo_alerta = resumo.copy()

    try:
        ret_path = escrever_excel_com_fallback(
            [out_path],
            resumo, evo,
            gerar_base, base_full,
            gerar_top, top_df
        )

        if deposito == "MAST" and prioridade_df is not None:
            with pd.ExcelWriter(ret_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as xw:
                prioridade_df.to_excel(xw, index=False, sheet_name="PRIORIDADE_OPERACIONAL_MAST")

        if deposito == "MAST" and not entradas_df.empty:
            with pd.ExcelWriter(ret_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as xw:
                entradas_df.to_excel(xw, index=False, sheet_name="ENTRADAS_MAST")
                if not entradas_dia_df.empty:
                    entradas_dia_df.to_excel(xw, index=False, sheet_name="ENTRADAS_MAST_DIA")

        ajustar_excel_visual(ret_path)
        atualizar_arquivo_atual(ret_path, out_dir, deposito)

        print_step(f"OUTPUT gerado (OFICIAL): {ret_path}")

    except PermissionError:
        resumo_alerta["AVISO"] = (
            "ATENÇÃO: Este relatório foi gerado como TEMP porque o arquivo oficial "
            "estava aberto no momento da atualização. "
            "Feche o Excel e execute novamente para gerar o relatório oficial."
        )

        candidatos = [temp_path] + [
            out_dir / f"{deposito}_resumo_{mref}__TEMP_{i:02d}.xlsx"
            for i in range(1, 21)
        ]

        try:
            ret_path = escrever_excel_com_fallback(
                candidatos,
                resumo_alerta, evo,
                gerar_base, base_full,
                gerar_top, top_df
            )

            if deposito == "MAST" and prioridade_df is not None:
                with pd.ExcelWriter(ret_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as xw:
                    prioridade_df.to_excel(xw, index=False, sheet_name="PRIORIDADE_OPERACIONAL_MAST")

            if deposito == "MAST" and not entradas_df.empty:
                with pd.ExcelWriter(ret_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as xw:
                    entradas_df.to_excel(xw, index=False, sheet_name="ENTRADAS_MAST")
                    if not entradas_dia_df.empty:
                        entradas_dia_df.to_excel(xw, index=False, sheet_name="ENTRADAS_MAST_DIA")

            ajustar_excel_visual(ret_path)
            atualizar_arquivo_atual(ret_path, out_dir, deposito)

            print_step(f"OUTPUT gerado (TEMP): {ret_path}")

            aviso_path = out_dir / f"AVISO_ARQUIVO_BLOQUEADO_{deposito}_{mref}.txt"
            aviso_msg = (
                "ATENÇÃO – ARQUIVO DE OUTPUT BLOQUEADO\n\n"
                "O relatório oficial não pôde ser atualizado porque estava aberto\n"
                "(em uso) no momento da execução.\n\n"
                f"Data/Hora: {now_ts()}\n"
                f"Depósito: {deposito}\n\n"
                "Arquivo oficial:\n"
                f"{out_path}\n\n"
                "Arquivo gerado como alternativa (TEMP):\n"
                f"{ret_path}\n\n"
                "O que fazer agora:\n"
                "1) Feche o arquivo Excel que estiver aberto.\n"
                "2) Execute novamente a atualização para gerar o relatório oficial.\n"
                "3) Enquanto isso, utilize o arquivo TEMP para consulta.\n\n"
            )
            aviso_path.write_text(aviso_msg, encoding="utf-8")

        except PermissionError:
            aviso_path = out_dir / f"AVISO_FALHA_GRAVACAO_{deposito}_{mref}.txt"
            aviso_msg = (
                "FALHA AO GRAVAR OUTPUT\n"
                f"Data/Hora: {now_ts()}\n"
                f"Depósito: {deposito}\n"
                f"Tentou oficial: {out_path}\n"
                f"Tentou temp: {temp_path} e TEMP_01..TEMP_20\n\n"
                "Causa provável:\n"
                "OFICIAL e TEMP estão abertos/bloqueados no Excel.\n\n"
                "Ação:\n"
                "Feche os arquivos e rode novamente.\n"
            )
            aviso_path.write_text(aviso_msg, encoding="utf-8")
            raise RuntimeError(
                f"Não foi possível gravar o OUTPUT: OFICIAL e TEMP estão bloqueados. Veja {aviso_path.name}"
            )

    print_step(f"BASE_CONSUMO linhas: {len(base_full)} | TOP linhas: {len(top_df)}")
    return ret_path


def run(config_path, snapshot_date=None):
    """
    Executa o ETL legado (V3) sem argparse (chamável via pipeline.py).

    Entradas:
    - config_path: caminho do config.json (pode ser relativo/absoluto, mas geralmente é um path resolvido)
    - snapshot_date:
        * 'YYYY-MM-DD' para forçar data do snapshot
        * None para usar a data de hoje (date.today())

    Saídas (efeitos):
    - Atualiza/insere snapshots em fact_estoque_snapshot
    - Atualiza dimensões (processo/grupo/custo) quando os arquivos existem
    - Atualiza/insere MB51 (entradas MAST) quando existe arquivo e há dados após filtros
    - Garante baseline do mês (primeiro dia útil) por depósito
    - Calcula e salva KPI mensal por depósito
    - Gera outputs em OUTPUT/<DEP>/...
    - Escreve um log resumido em LOGS/<YYYY-MM-DD>.log
    """

    # Log explícito (console) para o usuário e para rastreabilidade
    print(f"[update_from_sap] run chamado | config_path={config_path} | snapshot_date={snapshot_date}")

    # ------------------------------------------------------------
    # 1) Carregamento do config.json e definição do project_root
    # ------------------------------------------------------------
    cfg_path = Path(config_path)

    # Lê o JSON de config (utf-8)
    cfg = json.loads(cfg_path.read_text(encoding="utf-8"))

    # project_root:
    # - pode vir no config, senão assume a pasta onde está o config.json
    project_root = Path(cfg.get("project_root", cfg_path.parent)).resolve()

    # ------------------------------------------------------------
    # 2) Garante estrutura de pastas padrão do projeto
    # ------------------------------------------------------------
    # Essas pastas são usadas para:
    # - SAP_IN: entrada do snapshot de estoque
    # - DB: arquivo SQLite

    # Pastas (preferência: caminhos explícitos do config; fallback: padrão no project_root)
    sap_in_cfg = cfg.get("sap_in_dir", project_root / "SAP_IN")
    mb51_in_cfg = cfg.get("mb51_in_dir", project_root / "MB51_IN")
    db_cfg = cfg.get("db_path", project_root / "DB" / "estoque.sqlite")
    out_root_cfg = cfg.get("output_dir", project_root / "OUTPUT")
    log_root_cfg = cfg.get("logs_dir", project_root / "LOGS")

    sap_in = Path(sap_in_cfg)
    mb51_in = Path(mb51_in_cfg)
    db_path = Path(db_cfg)
    out_root = Path(out_root_cfg)
    log_dir = Path(log_root_cfg)

    # Resolve relativos contra project_root
    if not sap_in.is_absolute():
        sap_in = (project_root / sap_in).resolve()
    if not mb51_in.is_absolute():
        mb51_in = (project_root / mb51_in).resolve()
    if not db_path.is_absolute():
        db_path = (project_root / db_path).resolve()
    if not out_root.is_absolute():
        out_root = (project_root / out_root).resolve()
    if not log_dir.is_absolute():
        log_dir = (project_root / log_dir).resolve()

    # Garante diretórios reais
    sap_in.mkdir(parents=True, exist_ok=True)
    mb51_in.mkdir(parents=True, exist_ok=True)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    out_root.mkdir(parents=True, exist_ok=True)
    log_dir.mkdir(parents=True, exist_ok=True)

    # Pastas de output por depósito
    for dep in ("MAST", "MASR", "WEPV"):
        (out_root / dep).mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------
    # 3) Resolve caminhos (db_path, sap_in_dir e mb51_in_dir)
    # ------------------------------------------------------------
    # db_path pode vir no config como:
    # - absoluto (ex.: \\server\...\DB\estoque.sqlite)
    # - relativo (ex.: DB/estoque.sqlite)
    # Regra: se for relativo, resolve contra project_root
    db_cfg = cfg.get("db_path", project_root / "DB" / "estoque.sqlite")
    db_path = Path(db_cfg)
    if not db_path.is_absolute():
        db_path = (project_root / db_path).resolve()
    else:
        db_path = db_path.resolve()

    # sap_in_dir (onde está o arquivo principal de snapshot SAP)
    sap_in_cfg = cfg.get("sap_in_dir", project_root / "SAP_IN")
    sap_in = Path(sap_in_cfg)
    if not sap_in.is_absolute():
        sap_in = (project_root / sap_in).resolve()
    else:
        sap_in = sap_in.resolve()

    # ============================================================
    # 4) MB51 (entradas MAST) — resolve caminho corretamente (V3)
    # ============================================================
    # mb51_in_dir:
    # - pasta onde ficam os arquivos MB51MAST*.xlsx
    # mb51_sheet_name:
    # - nome da aba esperada (padrão "Data")
    mb51_in_cfg = cfg.get("mb51_in_dir", "MB51_IN")
    mb51_sheet_name = cfg.get("mb51_sheet_name", "Data")
    mb51_file = None

    mb51_in = Path(mb51_in_cfg)
    if not mb51_in.is_absolute():
        mb51_in = (project_root / mb51_in).resolve()
    else:
        mb51_in = mb51_in.resolve()

    print_step(f"MB51: procurando pasta em: {mb51_in}")

    # ------------------------------------------------------------
    # 5) Fontes auxiliares da nova arquitetura
    # ------------------------------------------------------------
    # Base Família Comercial:
    # - leitura direta do CSV oficial do BI.
    #
    # Processamento:
    # - consulta direta ao ASSIST, executada por carregar_dim_material().
    #
    # Hierarquia BU/Diretoria/Centro de Custo:
    # - consumida da tabela dim_segmento já homologada no Banco Estoques.
    familia_csv = Path(cfg.get("familia_csv", ""))

    if not familia_csv.is_absolute():
        familia_csv = (project_root / familia_csv).resolve()
    else:
        familia_csv = familia_csv.resolve()

    if not familia_csv.exists():
        raise FileNotFoundError(
            "CSV oficial da Base Família Comercial não encontrado: "
            f"{familia_csv}"
        )

    custo_file = Path(cfg.get("custo_file", "")).resolve()
    custo_sheet = cfg.get("custo_sheet", "Sheet1")
    custo_col_material = cfg.get("custo_col_material", "Material")
    custo_col_custo = cfg.get("custo_col_custo", "Custo Material")

    # Meta do KPI (padrão 1.2 se não vier no config)
    meta_giro = float(cfg.get("meta", {}).get("meta_giro_valor", 1.2))

    # ------------------------------------------------------------
    # 6) Definição da data do snapshot (sd = YYYY-MM-DD)
    # ------------------------------------------------------------
    if snapshot_date:
        d = datetime.strptime(snapshot_date, "%Y-%m-%d").date()
    else:
        d = date.today()
    sd = to_date_str(d)

    print_step(f"Início update snapshot={sd} (gera OUTPUT: MAST + MASR + WEPV)")

    # ------------------------------------------------------------
    # 7) Leitura do SAP principal (snapshot de estoque)
    # ------------------------------------------------------------
    # pick_latest_xlsx:
    # - escolhe o arquivo .xlsx mais recente na pasta SAP_IN
    sap_file = pick_latest_xlsx(sap_in)
    print_step(f"SAP: usando arquivo {sap_file.name}")

    # load_sap:
    # - carrega e normaliza o relatório do SAP (snapshot)
    sap_df = load_sap(sap_file)

    # ============================================================
    # 8) MB51: carrega o arquivo mais recente (se existir)
    # ============================================================
    # Observação importante de negócio:
    # - o arquivo do dia pode existir, mas não conter movimentação naquele dia
    # - por isso depois filtramos e pode acabar gravando 0 linhas (comportamento OK)
    mb51_df = pd.DataFrame()
    if mb51_in.exists():
        try:
            mb51_file = pick_latest_xlsx(mb51_in)
            print_step(f"MB51: usando arquivo {mb51_file.name}")
            mb51_df = load_mb51(mb51_file, sheet_name=mb51_sheet_name)
            print_step(f"MB51: linhas lidas do Excel: {len(mb51_df)}")
            
            # LOG MB51 (Excel): faixa de datas e dias distintos (antes dos filtros)
            if (not mb51_df.empty) and ("data_lancamento" in mb51_df.columns):
                _d = pd.to_datetime(mb51_df["data_lancamento"], errors="coerce").dropna()
                if not _d.empty:
                    print_step(
                        "MB51 (Excel): "
                        f"data_min={_d.min().date()} | data_max={_d.max().date()} | "
                        f"dias_distintos={_d.dt.date.nunique()}"
                    )
                else:
                    print_step("MB51 (Excel): coluna data_lancamento sem datas válidas (após parse).")
            else:
                print_step("MB51 (Excel): sem dados ou sem coluna data_lancamento (antes dos filtros).")

        except FileNotFoundError:
            print_step("MB51: nenhum arquivo encontrado em MB51_IN (ok, segue sem)")
    else:
        print_step("MB51: pasta MB51_IN não encontrada (ok, segue sem)")


    # ------------------------------------------------------------
    # 9) Atualização da dimensão consolidada de materiais
    # ------------------------------------------------------------
    # A dim_segmento já deve existir no Banco Estoques e é administrada
    # por uma carga controlada separada.
    #
    # Esta etapa atualiza dim_material usando:
    # - CSV oficial do BI;
    # - ASSIST;
    # - dim_segmento.
    print_step("DIM_MATERIAL: iniciando carga consolidada")

    carregar_dim_material(
        db_path=db_path,
        familia_csv=familia_csv,
    )

    print_step("DIM_MATERIAL: carga consolidada concluída")

    # ------------------------------------------------------------
    # 10) Custo (custo unitário) — opcional
    # ------------------------------------------------------------
    custo_df = pd.DataFrame(columns=["material", "custo_unit"])
    custo_last = ""
    if custo_file.exists():
        custo_last = datetime.fromtimestamp(custo_file.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        custo_df = load_custo_base(custo_file, custo_sheet, custo_col_material, custo_col_custo)
        print_step(f"CUSTO: {custo_file.name} linhas={len(custo_df)}")
    else:
        print_step("CUSTO: arquivo não encontrado (ok, segue sem)")

    # ------------------------------------------------------------
    # 12) Conexão com o SQLite e flags de execução
    # ------------------------------------------------------------
    con = sqlite3.connect(str(db_path))

    # baseline_created:
    # - flag global só para log final (se algum depósito criou baseline hoje)
    baseline_created = False

    # outputs:
    # - guarda caminhos dos outputs gerados por depósito para print/log final
    outputs = {}

    try:
        # Ativa integridade referencial caso existam FKs
        con.execute("PRAGMA foreign_keys=ON;")

        # ------------------------------------------------------------
        # 13) Garantia de schema (tabelas e colunas necessárias)
        # ------------------------------------------------------------
        # Aqui garantimos que todas as estruturas existam antes de gravar.
        ensure_fact_snapshot(con)
        ensure_baseline_tables(con)
        ensure_kpi_table(con)
        ensure_kpi_diario_table(con)
        # dim_material já foi atualizada antes da abertura desta conexão.
        # As dimensões legadas permanecem no banco durante a transição,
        # mas não são mais alimentadas por este ETL.
        ensure_dim_material_custo(con)
        ensure_mb51_schema(con)

        # 14) Snapshot (fato principal) — grava no banco

        # 🔴 CORREÇÃO CRÍTICA: garante snapshot "limpo" do dia
        con.execute("""
            DELETE FROM fact_estoque_snapshot
            WHERE snapshot_date = ?
        """, (sd,))

        con.commit()

        print_step(f"SNAPSHOT LIMPO para {sd} (remoção antes de regravar)")

        # Agora grava o snapshot correto
        upsert_snapshot(con, sap_df, sd, sap_file.name)

        # ------------------------------------------------------------
        # 15) Dimensões auxiliares
        # ------------------------------------------------------------
        # dim_material foi carregada antes da abertura desta conexão.
        # Nesta etapa permanece apenas a atualização da dimensão de custo.
        if not custo_df.empty:
            upsert_dim_material_custo(
                con,
                custo_df,
                custo_file.name,
                custo_last,
            )

        # ------------------------------------------------------------
        # 16) MB51 — carga e gravação (MAST)  [PASSO 1: fazer rodar]
        # ------------------------------------------------------------
        if not mb51_df.empty:
            mb = mb51_df.copy()

            # Depósito alvo
            mb = mb[mb["deposito"].astype("string").str.upper().str.strip() == "MAST"].copy()

            mb["quantidade"] = pd.to_numeric(mb["quantidade"], errors="coerce").fillna(0)

            # MB51: saídas (H) normalmente vêm com quantidade NEGATIVA.
            # A direção (entrada/saída) já está em deb_cred (S/H), então aqui guardamos apenas a magnitude.
            mb["quantidade"] = mb["quantidade"].abs()

            # Mantém apenas movimentos com quantidade diferente de zero
            mb = mb[mb["quantidade"] > 0].copy()


            # Normalizações
            mb["ordem"] = mb["ordem"].fillna("").astype("string").str.strip()
            mb["deb_cred"] = mb["deb_cred"].fillna("").astype("string").str.strip().str.upper()

            # PASSO 1 (correção): NÃO filtra aqui. Grava o MB51 completo.
            mb_mov = mb.copy()

            # Log de auditoria (sem variáveis inexistentes)
            entradas_validas = int(((mb_mov["deb_cred"] == "S") & (mb_mov["ordem"] == "")).sum())
            saidas_validas   = int(((mb_mov["deb_cred"] == "H") & (mb_mov["ordem"] != "")).sum())
            total_s          = int((mb_mov["deb_cred"] == "S").sum())
            total_h          = int((mb_mov["deb_cred"] == "H").sum())

            ensure_fact_mb51_mov(con)
            upsert_mb51_mov(con, mb_mov, mb51_file.name if mb51_file else "MB51.xlsx")

            print_step(
                "MB51: gravado COMPLETO (sem filtro na carga) | "
                f"linhas={len(mb_mov)} | S={total_s} | H={total_h} | "
                f"entradas_validas(S sem OP)={entradas_validas} | "
                f"saidas_validas(H com OP)={saidas_validas}"
            )
        else:
            print_step("MB51: sem dados (ok, segue sem)")


        # ------------------------------------------------------------
        # 17) Processamento por depósito (baseline + KPI + output)
        # ------------------------------------------------------------
        for dep in ["MAST", "MASR", "WEPV"]:
            print_step(f"--- Processando depósito {dep} ---")

            # baseline:
            # - garante baseline do mês (primeiro dia útil)
            # - pode criar 1 vez por mês e reutilizar
            baseline_created = ensure_baseline(con, dep, d, sap_file.name) or baseline_created
            con.commit()

            # KPI mensal:
            # - calcula e grava em kpi_mensal_resumo
            compute_and_store_kpi(con, dep, d, meta_giro)
            con.commit()

            backfill_kpi_diario_mes(con, dep, d)
            con.commit()

            # Output:
            # - gera arquivos Excel/relatórios (dependendo da implementação)
            out_path = build_output(con, dep, d, out_root / dep, cfg)
            con.commit()

            outputs[dep] = str(out_path)

    finally:
        # Fecha conexão com DB sempre
        con.close()

    # ------------------------------------------------------------
    # 18) Log resumido diário (arquivo LOGS/<sd>.log)
    # ------------------------------------------------------------
    # Mesmo tendo logs detalhados no console, este arquivo serve como auditoria do dia.
    # log_dir já foi resolvido acima via config (logs_dir) com fallback
    log_file = log_dir / f"{sd}.log"

    with log_file.open("a", encoding="utf-8") as f:
        for dep in ["MAST", "MASR", "WEPV"]:
            out_name = Path(outputs.get(dep, "")).name if outputs.get(dep) else "NA"

            f.write(
                f"[{now_ts()}] OK snapshot={sd} dep={dep} sap={sap_file.name} "
                f"dim_material=OK familia_csv={familia_csv.name} assist=OK "
                f"custo={'OK' if custo_file.exists() else 'NA'} "
                f"baseline_created={baseline_created} output={out_name}\n"
            )

    # ------------------------------------------------------------
    # 19) Saída final no console
    # ------------------------------------------------------------
    print_step(f"OK: Snapshot {sd} atualizado.")
    for dep, p in outputs.items():
        print_step(f"OUTPUT {dep}: {p}")

# ============================================================
# Main
# ============================================================

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True, help="Caminho do config.json do projeto")
    ap.add_argument("--snapshot-date", default=None, help="Data do snapshot (YYYY-MM-DD). Se vazio, usa hoje.")
    args = ap.parse_args()

    run(args.config, snapshot_date=args.snapshot_date)

if __name__ == "__main__":
    main()
